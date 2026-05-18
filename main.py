"""小红书工具箱 - 多账号并行 + 结果面板"""
import json
import os
import platform as _platform
import queue
import random
import sys
import threading
import time
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, simpledialog, ttk

# 跨平台字体：Windows 用微软雅黑 / Mac 用 PingFang SC / 其余用系统默认
_SYS = _platform.system()
FONT_UI   = "PingFang SC"   if _SYS == "Darwin" else "Microsoft YaHei"
FONT_MONO = "Menlo"         if _SYS == "Darwin" else "Consolas"

from scraper import (ACCOUNTS_DIR, list_accounts, get_proxy, set_proxy,
                     parse_proxy, get_account_meta, fetch_proxy_pool)
from session import AccountSession
from exporter import export_search, export_note, export_intent_users
from analyzer import extract_intent_users, parse_xhs_count
import db
import ai
import license_mgr as lm
import activation_ui
import settings_mgr as sm
import scheduler as sch
import sign_updater
import announce_client
import proxy_pool as pp
import image_proc
import app_updater


class App:
    def __init__(self, root, license_data=None):
        self.root = root
        self.license_data = license_data or {}
        plan = lm.PLAN_NAMES.get(self.license_data.get("plan", ""),
                                  self.license_data.get("plan", ""))
        exp = self.license_data.get("expires_at", 0)
        remain = lm.fmt_remain(exp) if exp else ""
        root.title(f"绮绮采集器（仅供学习）v2.5  |  套餐: {plan}  |  剩余: {remain}")
        root.geometry("1320x980")
        root.minsize(1180, 760)
        root.protocol("WM_DELETE_WINDOW", self.on_close)
        # 设置窗口图标（Mac 不支持 .ico，只用 iconphoto）
        try:
            assets = Path(os.path.dirname(os.path.abspath(__file__))) / "assets"
            png = assets / "logo.png"
            if png.exists():
                self._icon_photo = tk.PhotoImage(file=str(png))
                root.iconphoto(True, self._icon_photo)
            # Windows 额外设置 .ico（Mac 上 iconbitmap 会报错，跳过）
            if _SYS == "Windows":
                ico = assets / "logo.ico"
                if ico.exists():
                    root.iconbitmap(default=str(ico))
        except Exception:
            pass

        self.sessions = {}
        self.log_q = queue.Queue()
        # 结果数据 (note_id -> row dict)
        self.results_data = {}
        # 排序状态：列名 -> 升序(True)/降序(False)
        self._sort_dir = {}

        if getattr(sys, "frozen", False):
            base = Path(os.path.dirname(sys.executable))
        else:
            base = Path(os.path.dirname(os.path.abspath(__file__)))
        self.out_dir = base / "output"
        self.out_dir.mkdir(exist_ok=True)

        self.acc_var = tk.StringVar()
        self._build_ui()
        self._build_menu()
        self._refresh_accounts()
        self._apply_settings()
        self._start_log_file()
        self._poll_log()
        self._poll_status()
        self._auto_save_settings()
        self._check_license_expiry()
        # 启动定时任务调度器
        self.sched = sch.Scheduler(
            runner_cb=self._scheduled_runner,
            log_cb=lambda m: self.log("SCHED", m),
        )
        self.sched.start()
        # 异步检查签名 JS 更新
        threading.Thread(target=self._check_sign_update, daemon=True).start()
        # 异步拉取公告
        threading.Thread(target=self._fetch_announces_on_start, daemon=True).start()
        # 异步检查软件版本
        threading.Thread(target=self._check_app_update_silent, daemon=True).start()
        self.log("SYS", f"输出目录: {self.out_dir}")
        self.log("SYS", "搜索/爆品结果会显示在中间面板，可勾选后执行批量操作")
        self.log("SYS", f"MySQL: {'✓ 已配置' if db.is_enabled() else '✗ 未配置（去重/历史不可用 - 菜单 ⚙→MySQL）'}")
        src = ai.current_source()
        ai_status = {
            "user": "✓ 已配置（用户自定义）",
            "server": "✓ 已配置（系统默认 - 服务器下发）",
            "none": "✗ 未配置（AI 改写不可用 - 菜单 ⚙→DeepSeek）",
        }.get(src, "未知")
        self.log("SYS", f"DeepSeek: {ai_status}")
        # 启动心跳线程（每 5 分钟一次，上报使用 IP/版本/操作量）
        self._heartbeat_action_delta = 0
        threading.Thread(target=self._heartbeat_loop, daemon=True).start()

    # ============ UI 构建 ============
    # ─────────────────────────────────────────────────────────────────────
    # UI 构建辅助：带 bootstyle 回退的按钮工厂
    # ─────────────────────────────────────────────────────────────────────
    @staticmethod
    def _btn(parent, text, command, bs="secondary", width=None, **kw):
        try:
            import ttkbootstrap as _ttkb
            b = _ttkb.Button(parent, text=text, command=command,
                             bootstyle=bs, **kw)
        except Exception:
            b = ttk.Button(parent, text=text, command=command, **kw)
        if width:
            b.configure(width=width)
        return b

    @staticmethod
    def _card(parent, title="", padding=10):
        """现代卡片容器：顶部彩色标题条 + 内容区"""
        outer = ttk.Frame(parent, style="Card.TFrame")
        if title:
            hdr = ttk.Label(outer, text=title,
                            font=(FONT_UI, 10, "bold"),
                            style="CardHeader.TLabel")
            hdr.pack(fill="x", ipady=5, ipadx=10)
        inner = ttk.Frame(outer, padding=padding)
        inner.pack(fill="both", expand=True)
        return outer, inner

    def _build_ui(self):
        # ══════════════════════════════════════════════════════
        # 1. 顶部 Header Bar
        # ══════════════════════════════════════════════════════
        try:
            import ttkbootstrap as ttkb
            hdr = ttkb.Frame(self.root, bootstyle="primary", padding=(14, 8))
        except Exception:
            hdr = tk.Frame(self.root, bg="#F25928", pady=8)
        hdr.pack(fill="x", side="top")

        # 左侧：logo 文字 + 版本
        try:
            import ttkbootstrap as ttkb
            ttkb.Label(hdr, text="绮绮采集器",
                       font=(FONT_UI, 15, "bold"),
                       bootstyle="inverse-primary").pack(side="left")
            ttkb.Label(hdr, text="  v2.8  XHS Scraper Pro",
                       font=(FONT_UI, 9),
                       bootstyle="inverse-primary",
                       foreground="#ffffff99").pack(side="left", pady=(3, 0))
        except Exception:
            tk.Label(hdr, text="绮绮采集器  v2.8",
                     bg="#F25928", fg="white",
                     font=(FONT_UI, 14, "bold")).pack(side="left")

        # 右侧：套餐 + 剩余 + 日志按钮
        plan = lm.PLAN_NAMES.get(self.license_data.get("plan", ""),
                                  self.license_data.get("plan", "—"))
        exp   = self.license_data.get("expires_at", 0)
        remain = lm.fmt_remain(exp) if exp else "—"
        try:
            import ttkbootstrap as ttkb
            ttkb.Label(hdr, text=f"套餐：{plan}",
                       bootstyle="inverse-primary",
                       font=(FONT_UI, 10)).pack(side="right", padx=(0, 6))
            ttkb.Label(hdr, text=f"剩余：{remain}",
                       bootstyle="inverse-primary",
                       font=(FONT_UI, 10)).pack(side="right", padx=(0, 16))
            self._btn(hdr, "📋 日志", self._show_log_window,
                      bs="secondary-outline").pack(side="right", padx=4)
        except Exception:
            tk.Label(hdr, text=f"套餐:{plan}  剩余:{remain}",
                     bg="#F25928", fg="white",
                     font=(FONT_UI, 10)).pack(side="right", padx=12)
            ttk.Button(hdr, text="📋 日志",
                       command=self._show_log_window).pack(side="right", padx=4)

        # ══════════════════════════════════════════════════════
        # 2. 主体：左侧边栏 + 右侧内容
        # ══════════════════════════════════════════════════════
        body = ttk.Frame(self.root)
        body.pack(fill="both", expand=True, side="top")

        # ── 左侧边栏 ──────────────────────────────────────────
        sidebar = ttk.Frame(body, width=220)
        sidebar.pack(side="left", fill="y", padx=(6, 0), pady=6)
        sidebar.pack_propagate(False)   # 固定 220px 宽

        # 账号列表卡片
        acc_hdr = ttk.Label(sidebar, text="👤  账号管理",
                             font=(FONT_UI, 10, "bold"))
        acc_hdr.pack(fill="x", pady=(0, 4))
        ttk.Separator(sidebar, orient="horizontal").pack(fill="x", pady=(0, 6))

        tv_wrap = ttk.Frame(sidebar)
        tv_wrap.pack(fill="both", expand=True)
        cols = ("alias", "nick", "proxy", "login", "status")
        self.tv = ttk.Treeview(tv_wrap, columns=cols, show="headings",
                               height=6, selectmode="browse")
        self.tv.heading("alias",  text="账号")
        self.tv.heading("nick",   text="昵称")
        self.tv.heading("proxy",  text="代理")
        self.tv.heading("login",  text="登录")
        self.tv.heading("status", text="状态")
        self.tv.column("alias",  width=70,  anchor="w", stretch=False)
        self.tv.column("nick",   width=0,   stretch=False)   # 隐藏（保留兼容）
        self.tv.column("proxy",  width=0,   stretch=False)
        self.tv.column("login",  width=0,   stretch=False)
        self.tv.column("status", width=140, anchor="w")
        self.tv.pack(side="left", fill="both", expand=True)
        sb_acc = ttk.Scrollbar(tv_wrap, orient="vertical", command=self.tv.yview)
        sb_acc.pack(side="right", fill="y")
        self.tv.configure(yscrollcommand=sb_acc.set)
        self.tv.bind("<<TreeviewSelect>>", self.on_tv_select)
        self.tv.bind("<Double-1>", lambda e: self.on_login())

        # 账号操作按钮区
        ttk.Separator(sidebar, orient="horizontal").pack(fill="x", pady=8)
        btn_grid = ttk.Frame(sidebar)
        btn_grid.pack(fill="x")
        # 2×2 格局
        self._btn(btn_grid, "+ 新账号",    self.on_acc_new,      bs="primary",
                  width=13).grid(row=0, column=0, padx=2, pady=2, sticky="ew")
        self._btn(btn_grid, "🌐 代理",     self.on_proxy_edit,   bs="info-outline",
                  width=10).grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        self._btn(btn_grid, "▶ 登录",      self.on_login,        bs="success",
                  width=13).grid(row=1, column=0, padx=2, pady=2, sticky="ew")
        self._btn(btn_grid, "▶▶ 全部",    self.on_start_all,    bs="success-outline",
                  width=10).grid(row=1, column=1, padx=2, pady=2, sticky="ew")
        self._btn(btn_grid, "⏹ 停该号",   self.on_stop_selected, bs="danger-outline",
                  width=13).grid(row=2, column=0, padx=2, pady=2, sticky="ew")
        self._btn(btn_grid, "⏹ 全停",     self.on_stop_all,     bs="danger",
                  width=10).grid(row=2, column=1, padx=2, pady=2, sticky="ew")
        self._btn(btn_grid, "🔍 检测IP",   self.on_check_ip,     bs="secondary",
                  width=13).grid(row=3, column=0, padx=2, pady=2, sticky="ew")
        self._btn(btn_grid, "✕ 关窗口",   self.on_close_selected, bs="secondary-outline",
                  width=10).grid(row=3, column=1, padx=2, pady=2, sticky="ew")
        btn_grid.columnconfigure(0, weight=1)
        btn_grid.columnconfigure(1, weight=1)

        # 安全时间窗 + API 模式（折叠在底部）
        ttk.Separator(sidebar, orient="horizontal").pack(fill="x", pady=(10, 6))
        ttk.Label(sidebar, text="⏰  运行设置",
                  font=(FONT_UI, 10, "bold")).pack(anchor="w")
        tw_inner = ttk.Frame(sidebar, padding=(0, 4))
        tw_inner.pack(fill="x")
        ttk.Label(tw_inner, text="时间窗:").grid(row=0, column=0, sticky="w")
        self.e_hstart = ttk.Entry(tw_inner, width=4)
        self.e_hstart.insert(0, "10")
        self.e_hstart.grid(row=0, column=1, padx=2)
        ttk.Label(tw_inner, text="~").grid(row=0, column=2)
        self.e_hend = ttk.Entry(tw_inner, width=4)
        self.e_hend.insert(0, "23")
        self.e_hend.grid(row=0, column=3, padx=2)
        self.var_api_mode = tk.BooleanVar(value=True)
        ttk.Checkbutton(sidebar, text="⚡ API 直发模式",
                        variable=self.var_api_mode).pack(anchor="w", pady=(4, 0))
        ttk.Label(sidebar,
                  text="（不打开页面，速度 5-10×）",
                  foreground="#888",
                  font=(FONT_UI, 9)).pack(anchor="w")

        # ── 右侧主内容区 ──────────────────────────────────────
        main_area = ttk.Frame(body)
        main_area.pack(side="left", fill="both", expand=True,
                       padx=6, pady=6)

        # 进度条行（平时隐藏，运行时显示）
        self.prog_frame = ttk.Frame(main_area)
        self.prog_var = tk.IntVar(value=0)
        self.prog_label = ttk.Label(self.prog_frame, text="",
                                    foreground="#0a7", font=(FONT_UI, 10))
        self.prog_label.pack(side="left", padx=(0, 8))
        self.prog_bar = ttk.Progressbar(self.prog_frame,
                                         variable=self.prog_var,
                                         mode="determinate")
        self.prog_bar.pack(side="left", fill="x", expand=True)
        # 不 pack，由 _set_progress 控制

        # PanedWindow：上方 Tabs + 下方结果
        self.paned = ttk.PanedWindow(main_area, orient="vertical")
        self.paned.pack(fill="both", expand=True)

        nb_frame = ttk.Frame(self.paned)
        self.paned.add(nb_frame, weight=2)
        nb = ttk.Notebook(nb_frame)
        nb.pack(fill="both", expand=True)
        self._build_tabs(nb)

        results_frame = ttk.Frame(self.paned)
        self.paned.add(results_frame, weight=4)
        self._build_results(results_frame)

        # ══════════════════════════════════════════════════════
        # 3. 底部状态栏
        # ══════════════════════════════════════════════════════
        try:
            import ttkbootstrap as ttkb
            sbar = ttkb.Frame(self.root, bootstyle="secondary", padding=(10, 3))
        except Exception:
            sbar = tk.Frame(self.root, bg="#e8eaed", pady=3)
        sbar.pack(fill="x", side="bottom")
        self._status_var = tk.StringVar(value="就绪")
        try:
            import ttkbootstrap as ttkb
            self._status_lbl = ttkb.Label(sbar, textvariable=self._status_var,
                                          bootstyle="inverse-secondary",
                                          font=(FONT_UI, 9))
        except Exception:
            self._status_lbl = ttk.Label(sbar, textvariable=self._status_var,
                                          font=(FONT_UI, 9))
        self._status_lbl.pack(side="left")

        # ══════════════════════════════════════════════════════
        # 4. 日志独立浮窗
        # ══════════════════════════════════════════════════════
        self._build_log_window()

    # ---------- 日志浮窗 ----------
    def _build_log_window(self):
        self.log_win = tk.Toplevel(self.root)
        self.log_win.title("📋 运行日志")
        self.log_win.geometry("520x600")
        self.log_win.configure(bg="#1a1f26")
        # 关闭 X = 隐藏（保留内容，避免 self.txt 失效）
        self.log_win.protocol("WM_DELETE_WINDOW", self.log_win.withdraw)
        # 图标跟随主窗口（Mac 不支持 iconbitmap）
        if _SYS == "Windows":
            try:
                assets = Path(os.path.dirname(os.path.abspath(__file__))) / "assets"
                ico = assets / "logo.ico"
                if ico.exists():
                    self.log_win.iconbitmap(str(ico))
            except Exception:
                pass

        bar = ttk.Frame(self.log_win); bar.pack(fill="x", padx=6, pady=(6, 2))
        ttk.Label(bar, text="📋 运行日志",
                  font=(FONT_UI, 11, "bold")).pack(side="left")
        ttk.Button(bar, text="打开输出目录",
                   command=self.open_out).pack(side="right", padx=2)
        ttk.Button(bar, text="清空",
                   command=lambda: self.txt.delete("1.0", "end")).pack(side="right", padx=2)

        txt_wrap = ttk.Frame(self.log_win); txt_wrap.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        self.txt = tk.Text(txt_wrap, bg="#1a1f26", fg="#d4d4d4",
                           insertbackground="#fff",
                           font=(FONT_MONO, 11), padx=6, pady=4)
        sb = ttk.Scrollbar(txt_wrap, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.txt.pack(side="left", fill="both", expand=True)

        # 等主窗口渲染完再贴边
        self.root.after(200, self._position_log_window)

    def _position_log_window(self):
        """ 将日志窗口贴到主窗口右侧。右侧放不下就退到左侧或屏内最右 """
        try:
            self.root.update_idletasks()
            mx = self.root.winfo_x()
            my = self.root.winfo_y()
            mw = self.root.winfo_width()
            mh = self.root.winfo_height()
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            win_w = 520
            win_h = max(400, min(mh, sh - 80))
            x = mx + mw + 10
            if x + win_w > sw:
                # 右侧装不下 → 优先放到主窗口左侧
                alt_x = mx - win_w - 10
                x = alt_x if alt_x >= 0 else max(0, sw - win_w - 5)
            y = max(0, min(my, sh - win_h - 50))
            self.log_win.geometry(f"{win_w}x{win_h}+{x}+{y}")
        except Exception:
            pass

    def _show_log_window(self):
        """ 把日志窗口呼回前台并重新贴边 """
        try:
            self.log_win.deiconify()
            self.log_win.lift()
            self._position_log_window()
        except Exception:
            pass

    # ---------- 进度条 ----------
    def _set_progress(self, current, total, msg=""):
        """ 跨线程安全更新进度条 """
        def upd():
            if total <= 0:
                self.prog_frame.pack_forget()
                self.prog_var.set(0)
                self.prog_label.config(text="")
                return
            self.prog_var.set(int(current * 100 / max(total, 1)))
            self.prog_label.config(text=f"{msg}  {current}/{total}")
            if not self.prog_frame.winfo_ismapped():
                # 插在主 PanedWindow 之前（即时间窗与 Tabs 之间）
                self.prog_frame.pack(fill="x", before=self.paned)
        self.root.after(0, upd)

    # ---------- 设置持久化 ----------
    def _settings_widget_map(self):
        """ 返回 {key: (type, getter, setter)} """
        # type: 'text' 'entry' 'bool' 'str'
        return {
            "templates":      ("text",  self.txt_templates),
            "reply_templates":("text",  self.txt_reply),
            "search_keyword": ("entry", self.e_kw),
            "search_count":   ("entry", self.e_count),
            "note_input":     ("entry", self.e_note),
            "comment_count":  ("entry", self.e_cmt),
            "user_input":     ("entry", self.e_user),
            "user_count":     ("entry", self.e_ucount),
            "comment_min":    ("entry", self.e_dmin),
            "comment_max":    ("entry", self.e_dmax),
            "daily_limit":    ("entry", self.e_daily),
            "follow_min":     ("entry", self.e_fmin),
            "follow_max":     ("entry", self.e_fmax),
            "follow_limit":   ("entry", self.e_flim),
            "hour_start":     ("entry", self.e_hstart),
            "hour_end":       ("entry", self.e_hend),
            "hot_keyword":    ("entry", self.e_hkw),
            "hot_scan":       ("entry", self.e_hscan),
            "hot_min_like":   ("entry", self.e_hmin_like),
            "hot_min_cmt":    ("entry", self.e_hmin_cmt),
            "hot_min_col":    ("entry", self.e_hmin_col),
            "hot_top":        ("entry", self.e_htop),
            "hot_sort":       ("combo", self.cb_hsort),
            "dl_media":       ("bool",  self.var_dl_media),
            "extract_intent": ("bool",  self.var_extract_intent),
            "auto_reply":     ("bool",  self.var_auto_reply),
            "confirm":        ("bool",  self.var_confirm),
            "shuffle":        ("bool",  self.var_shuffle),
            "use_ai":         ("bool",  self.var_use_ai),
            "dedup":          ("bool",  self.var_dedup),
            "api_mode":       ("bool",  self.var_api_mode),
            "search_with_ip": ("bool",  self.var_search_with_ip),
        }

    def _apply_settings(self):
        """ 启动时把磁盘存的值灌进 UI """
        s = sm.load()
        for key, (kind, widget) in self._settings_widget_map().items():
            if key not in s:
                continue
            v = s[key]
            try:
                if kind == "text":
                    widget.delete("1.0", "end")
                    widget.insert("1.0", str(v))
                elif kind == "entry":
                    widget.delete(0, "end")
                    widget.insert(0, str(v))
                elif kind == "combo":
                    widget.set(str(v))
                elif kind == "bool":
                    widget.set(bool(v))
            except Exception:
                pass
        last = s.get("last_account", "")
        if last and last in list_accounts():
            self.acc_var.set(last)

    def _save_settings(self):
        """ 把当前 UI 状态全部存盘 """
        try:
            data = {}
            for key, (kind, widget) in self._settings_widget_map().items():
                try:
                    if kind == "text":
                        data[key] = widget.get("1.0", "end").rstrip("\n")
                    elif kind in ("entry", "combo"):
                        data[key] = widget.get()
                    elif kind == "bool":
                        data[key] = bool(widget.get())
                except Exception:
                    pass
            data["last_account"] = self.acc_var.get()
            sm.save_all(data)
        except Exception as e:
            self.log("SYS", f"保存设置失败: {e}")

    def _auto_save_settings(self):
        self._save_settings()
        self.root.after(60_000, self._auto_save_settings)

    # ---------- 文件日志 ----------
    def _start_log_file(self):
        try:
            log_dir = self.out_dir / "logs"
            log_dir.mkdir(exist_ok=True)
            self.log_file = open(
                log_dir / f"app_{time.strftime('%Y%m%d')}.log",
                "a", encoding="utf-8", buffering=1,
            )
            self.log_file.write(f"\n========== {time.strftime('%Y-%m-%d %H:%M:%S')} 启动 ==========\n")
        except Exception as e:
            self.log_file = None
            print(f"日志文件打开失败: {e}")

    # ---------- 授权过期提醒 ----------
    def _check_license_expiry(self):
        d = self.license_data or {}
        exp = d.get("expires_at", 0)
        if not exp:
            return
        remain_sec = int(exp) - int(time.time())
        if remain_sec < 0:
            return  # 已过期不会进到这
        days = remain_sec // 86400
        if days <= 3:
            self.log("SYS", f"⚠ 授权剩 {lm.fmt_remain(exp)}，请及时续费")

    def _heartbeat_loop(self):
        """ 每 5 分钟向服务器上报心跳：使用 IP / 操作量 / 版本
            如果服务器返回 blocked=True，强制退出软件 """
        while True:
            time.sleep(300)
            try:
                delta = self._heartbeat_action_delta
                self._heartbeat_action_delta = 0
                lm.heartbeat(action_delta=delta)
            except RuntimeError as e:
                # 被封禁
                self.log("SYS", f"🚫 {e}，软件将退出")
                self.root.after(0, lambda: self._force_quit(str(e)))
                return
            except Exception:
                pass

    def _force_quit(self, reason):
        try:
            messagebox.showerror("授权异常", f"{reason}\n请联系客服。")
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            pass
        os._exit(0)

    def plan_delay(self, override_min=None, override_max=None):
        """ 套餐级别的强制延时：客户端代码可以调这个拿到 (min, max) """
        limits = (self.license_data or {}).get("limits") or {}
        d_min = limits.get("action_delay_min", 30)
        d_max = limits.get("action_delay_max", 90)
        # 用户的本地设置不能低于服务端套餐下限（防止越级使用）
        if override_min is not None:
            d_min = max(d_min, int(override_min))
        if override_max is not None:
            d_max = max(d_max, int(override_max))
        return d_min, d_max

    def record_action(self, n=1):
        """ 记录用户操作次数（供心跳上报） """
        self._heartbeat_action_delta += n

    def _build_menu(self):
        bar = tk.Menu(self.root)
        self.root.config(menu=bar)

        # ⚙ 设置：配置类
        m_set = tk.Menu(bar, tearoff=0)
        bar.add_cascade(label="⚙ 设置", menu=m_set)
        m_set.add_command(label="🔑  授权信息", command=self.on_show_license)
        m_set.add_command(label="🗄  MySQL 数据库", command=self.on_db_config)
        m_set.add_command(label="🤖  DeepSeek AI", command=self.on_ai_config)
        m_set.add_command(label="🌐  代理 IP 池管理", command=self.on_proxy_pool)
        m_set.add_command(label="🖼  图片处理工具（水印/MD5扰动）", command=self.on_image_tool)
        m_set.add_separator()
        m_set.add_command(label="💾  立即保存设置", command=self._on_save_settings_btn)

        # 📊 数据
        m_data = tk.Menu(bar, tearoff=0)
        bar.add_cascade(label="📊 数据", menu=m_data)
        m_data.add_command(label="📊  今日操作统计", command=self.on_show_stats)
        m_data.add_command(label="📡  抓 API 请求", command=self.on_show_captured)

        # ⏰ 任务
        m_task = tk.Menu(bar, tearoff=0)
        bar.add_cascade(label="⏰ 任务", menu=m_task)
        m_task.add_command(label="⏰  定时任务管理", command=self.on_show_scheduled)

        # 📢 公告
        m_ann = tk.Menu(bar, tearoff=0)
        bar.add_cascade(label="📢 公告", menu=m_ann)
        m_ann.add_command(label="📢  查看最新公告", command=self.on_show_announces)
        m_ann.add_command(label="🆕  检查软件更新", command=self.on_check_app_update)
        m_ann.add_command(label="🔄  检查签名更新", command=self.on_check_sign_update_manual)

        # 🛡 安全
        m_sec = tk.Menu(bar, tearoff=0)
        bar.add_cascade(label="🛡 安全", menu=m_sec)
        m_sec.add_command(label="🛡  清除当前账号风控警报", command=self.on_clear_alert)
        m_sec.add_command(label="🔀  重置当前账号浏览器指纹", command=self.on_regen_fp)
        m_sec.add_command(label="🗑  删除当前账号", command=self.on_acc_delete)

        # 📦 备份
        m_bk = tk.Menu(bar, tearoff=0)
        bar.add_cascade(label="📦 备份", menu=m_bk)
        m_bk.add_command(label="📦  备份所有数据 → ZIP", command=self.on_backup)
        m_bk.add_command(label="📂  从 ZIP 恢复数据", command=self.on_restore)
        m_bk.add_separator()
        m_bk.add_command(label="📁  打开输出目录", command=self.open_out)

        # 🎨 主题
        m_theme = tk.Menu(bar, tearoff=0)
        bar.add_cascade(label="🎨 主题", menu=m_theme)
        m_light = tk.Menu(m_theme, tearoff=0)
        m_dark = tk.Menu(m_theme, tearoff=0)
        m_theme.add_cascade(label="☀ 浅色主题", menu=m_light)
        m_theme.add_cascade(label="🌙 深色主题", menu=m_dark)
        cur = getattr(self.root, "_current_theme", DEFAULT_THEME)
        for t in THEMES_LIGHT:
            mark = "  ●" if t == cur else "   "
            m_light.add_command(label=f"{mark}  {t}",
                                command=lambda name=t: self._on_switch_theme(name))
        for t in THEMES_DARK:
            mark = "  ●" if t == cur else "   "
            m_dark.add_command(label=f"{mark}  {t}",
                               command=lambda name=t: self._on_switch_theme(name))

    def _on_switch_theme(self, theme_name):
        """ 用户点击主题菜单 → 切换 + 持久化 + 重建菜单刷新勾选状态 """
        try:
            switch_theme(self.root, theme_name)
            self._build_menu()  # 重建以更新选中标记
            self.log("SYS", f"🎨 已切换主题: {theme_name}")
        except Exception as e:
            self.log("SYS", f"切换主题失败: {e}")

    # ---------- 数据库 / AI 配置弹窗 ----------
    def on_db_config(self):
        cfg = db.load_db_config()
        dlg = tk.Toplevel(self.root)
        dlg.title("MySQL 数据库配置")
        dlg.geometry("520x380")
        dlg.transient(self.root); dlg.grab_set()

        ttk.Label(dlg, text="用于：操作历史 / 去重 / 笔记沉淀 / 意向用户库",
                  foreground="#888").pack(anchor="w", padx=12, pady=(10, 6))
        ttk.Label(dlg, text="未配置时工具照常运行，只是不去重、不沉淀历史").pack(anchor="w", padx=12, pady=(0, 8))

        form = ttk.Frame(dlg); form.pack(padx=12, fill="x")
        fields = [("host", "主机", cfg.get("host", "127.0.0.1")),
                  ("port", "端口", str(cfg.get("port", 3306))),
                  ("user", "用户名", cfg.get("user", "root")),
                  ("password", "密码", cfg.get("password", "")),
                  ("database", "数据库名", cfg.get("database", "xhs_tool"))]
        entries = {}
        for i, (k, label, v) in enumerate(fields):
            ttk.Label(form, text=label).grid(row=i, column=0, sticky="e", pady=3, padx=4)
            e = ttk.Entry(form, width=40, show="*" if k == "password" else "")
            e.insert(0, v); e.grid(row=i, column=1, sticky="ew", pady=3)
            entries[k] = e
        form.columnconfigure(1, weight=1)

        msg = tk.StringVar()
        ttk.Label(dlg, textvariable=msg, foreground="#0a7",
                  wraplength=480, justify="left").pack(anchor="w", padx=12, pady=8)

        def collect():
            return {
                "host": entries["host"].get().strip(),
                "port": int(entries["port"].get() or "3306"),
                "user": entries["user"].get().strip(),
                "password": entries["password"].get(),
                "database": entries["database"].get().strip(),
            }

        def do_test():
            try:
                c = collect()
                v = db.test_connection(c)
                msg.set(f"✓ 连接成功，MySQL 版本: {v}")
            except Exception as e:
                msg.set(f"✗ {e}")

        def do_init():
            try:
                c = collect()
                db.save_db_config(c)
                db.init_schema()
                msg.set("✓ 表已创建（actions / notes / intent_users）")
            except Exception as e:
                msg.set(f"✗ {e}")

        def do_save():
            try:
                db.save_db_config(collect())
                msg.set("✓ 已保存")
                self.log("SYS", "MySQL 配置已保存")
            except Exception as e:
                msg.set(f"✗ {e}")

        b = ttk.Frame(dlg); b.pack(pady=8)
        ttk.Button(b, text="🧪 测试连接", command=do_test).pack(side="left", padx=4)
        ttk.Button(b, text="🔨 创建表", command=do_init).pack(side="left", padx=4)
        ttk.Button(b, text="💾 保存", command=do_save).pack(side="left", padx=4)
        ttk.Button(b, text="关闭", command=dlg.destroy).pack(side="left", padx=4)

    def on_ai_config(self):
        cfg = ai.load_ai_config()
        dlg = tk.Toplevel(self.root)
        dlg.title("DeepSeek AI 配置")
        dlg.geometry("760x720")
        dlg.minsize(680, 560)
        dlg.transient(self.root); dlg.grab_set()

        # ---- 底部按钮区先 pack（保证永远可见） ----
        msg = tk.StringVar()
        bottom = ttk.Frame(dlg); bottom.pack(side="bottom", fill="x", pady=8)
        ttk.Label(bottom, textvariable=msg, foreground="#0a7",
                  wraplength=720, justify="left").pack(side="top", padx=12, anchor="w")
        b = ttk.Frame(bottom); b.pack(pady=6)

        # ---- 中间用 Canvas + Scrollbar 包裹内容（防溢出） ----
        outer = ttk.Frame(dlg); outer.pack(side="top", fill="both", expand=True)
        canvas = tk.Canvas(outer, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = ttk.Frame(canvas)
        inner_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_cfg(e):
            canvas.itemconfigure(inner_id, width=e.width)
        inner.bind("<Configure>", _on_inner_cfg)
        canvas.bind("<Configure>", _on_canvas_cfg)
        # 鼠标滚轮支持
        def _on_wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # 真正的内容容器
        host = inner

        ttk.Label(host, text="DeepSeek API（兼容 OpenAI 格式）",
                  font=(FONT_UI, 11, "bold")).pack(anchor="w", padx=12, pady=(10, 4))
        ttk.Label(host, text="官网申请: https://platform.deepseek.com/  价格低于 GPT-4",
                  foreground="#888").pack(anchor="w", padx=12, pady=(0, 8))

        f = ttk.Frame(host); f.pack(fill="x", padx=12)
        ttk.Label(f, text="API Key:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        e_key = ttk.Entry(f, width=55, show="*"); e_key.insert(0, cfg.get("api_key", ""))
        e_key.grid(row=0, column=1, sticky="ew", pady=4)
        ttk.Label(f, text="模型:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
        cb_model = ttk.Combobox(f, values=["deepseek-chat", "deepseek-reasoner"],
                                 state="readonly", width=20)
        cb_model.set(cfg.get("model", "deepseek-chat"))
        cb_model.grid(row=1, column=1, sticky="w", pady=4)
        ttk.Label(f, text="Base URL:").grid(row=2, column=0, sticky="e", padx=4, pady=4)
        e_url = ttk.Entry(f, width=55)
        e_url.insert(0, cfg.get("base_url", "https://api.deepseek.com/v1/chat/completions"))
        e_url.grid(row=2, column=1, sticky="ew", pady=4)
        f.columnconfigure(1, weight=1)

        ttk.Label(host, text="System Prompt（决定 AI 的写作风格，可自定义）:").pack(anchor="w", padx=12, pady=(10, 2))
        txt_sys = tk.Text(host, height=8, width=72, font=(FONT_UI, 11))
        txt_sys.pack(padx=12, fill="x")
        txt_sys.insert("1.0", cfg.get("system_prompt", ai.DEFAULT_SYSTEM))

        ttk.Label(host, text="笔记生成设置（字数 / Prompt 模板）:").pack(anchor="w", padx=12, pady=(10, 2))
        f2 = ttk.Frame(host); f2.pack(fill="x", padx=12)
        ttk.Label(f2, text="图文/视频字数:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
        e_words_img = ttk.Entry(f2, width=18)
        e_words_img.insert(0, cfg.get("note_gen_words_image", "150-300"))
        e_words_img.grid(row=0, column=1, sticky="w", pady=4)

        ttk.Label(f2, text="长文字数:").grid(row=0, column=2, sticky="e", padx=8, pady=4)
        e_words_long = ttk.Entry(f2, width=18)
        e_words_long.insert(0, cfg.get("note_gen_words_longtext", "900-1500"))
        e_words_long.grid(row=0, column=3, sticky="w", pady=4)

        ttk.Label(f2, text="标签数量:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
        e_tags = ttk.Entry(f2, width=18)
        e_tags.insert(0, cfg.get("note_gen_tags", "5-8"))
        e_tags.grid(row=1, column=1, sticky="w", pady=4)
        f2.columnconfigure(3, weight=1)

        ttk.Label(host, text="笔记生成 System Prompt（可选，不填则用默认）:").pack(anchor="w", padx=12, pady=(8, 2))
        txt_note_sys = tk.Text(host, height=6, width=72, font=(FONT_UI, 11))
        txt_note_sys.pack(padx=12, fill="x")
        txt_note_sys.insert("1.0", cfg.get("note_gen_system", ai.NOTE_GEN_SYSTEM))

        ttk.Label(host, text="笔记生成 User Prompt 模板（占位符: {topic} {type_hint} {style} {words_min} {words_max} {tags_min} {tags_max}）:").pack(anchor="w", padx=12, pady=(8, 2))
        txt_note_tpl = tk.Text(host, height=6, width=72, font=(FONT_UI, 11))
        txt_note_tpl.pack(padx=12, fill="x", pady=(0, 12))
        txt_note_tpl.insert("1.0", cfg.get("note_gen_user_template", ai.DEFAULT_NOTE_USER_TEMPLATE))

        def collect():
            return {
                "api_key": e_key.get().strip(),
                "model": cb_model.get(),
                "base_url": e_url.get().strip(),
                "system_prompt": txt_sys.get("1.0", "end").strip(),
                "note_gen_words_image": e_words_img.get().strip(),
                "note_gen_words_longtext": e_words_long.get().strip(),
                "note_gen_tags": e_tags.get().strip(),
                "note_gen_system": txt_note_sys.get("1.0", "end").strip(),
                "note_gen_user_template": txt_note_tpl.get("1.0", "end").strip(),
            }

        def do_test():
            try:
                c = collect()
                ai.save_ai_config(c)
                msg.set("请求 DeepSeek...")
                dlg.update_idletasks()
                out = ai.rewrite("学到了 感谢分享",
                                 {"title": "5 分钟快手早餐", "type": "normal", "author": "饿货阿丽"})
                msg.set(f"✓ AI 返回: {out}")
            except Exception as e:
                msg.set(f"✗ {e}")

        def do_save():
            try:
                ai.save_ai_config(collect())
                msg.set("✓ 已保存")
                self.log("SYS", "DeepSeek 配置已保存")
            except Exception as e:
                msg.set(f"✗ {e}")

        ttk.Button(b, text="🧪 测试一次改写", command=do_test).pack(side="left", padx=4)
        ttk.Button(b, text="💾 保存", command=do_save).pack(side="left", padx=4)
        ttk.Button(b, text="关闭", command=dlg.destroy).pack(side="left", padx=4)

    # ---------- 定时任务 ----------
    def _scheduled_runner(self, task):
        """ 调度器线程触发 → 转换成对应账号的 session 任务 """
        acc = task.get("account")
        ttype = task.get("type")
        params = task.get("params") or {}
        if acc not in list_accounts():
            self.log("SCHED", f"账号 {acc} 不存在，跳过")
            return
        sess = self._get_session(acc, create_if_missing=True)
        if sess is None:
            return
        if ttype == "nurture":
            d = int(params.get("duration", 30))
            lp = float(params.get("like_prob", 0.15))
            cp = float(params.get("collect_prob", 0.05))
            sess.submit_automation(self._t_nurture, (d, lp, cp),
                                    f"⏰🌱 定时养号 {d}min")
        elif ttype == "search_like":
            kw = params.get("keyword", "")
            count = int(params.get("count", 20))
            min_like = int(params.get("min_like", 1000))
            # 简单实现: 搜索 + 自动点赞所有结果
            def task(scraper, kw=kw, count=count, min_like=min_like):
                rows = scraper.search_hot_notes(kw, max_scan=count, min_liked=min_like,
                                                 top_n=count, sort_by="liked_count")
                if not rows: return
                self._t_bulk_like(scraper, rows, 30, 90,
                                   self.var_dedup.get(), True)
            sess.submit_automation(task, (), f"⏰🔍 定时点赞「{kw}」")
        elif ttype == "follow_authors":
            kw = params.get("keyword", "")
            count = int(params.get("count", 10))
            def task(scraper, kw=kw, count=count):
                rows = scraper.search_notes(kw, max_count=count)
                items = []
                for r in rows:
                    aid = r.get("author_id")
                    if aid:
                        items.append((aid, f"https://www.xiaohongshu.com/user/profile/{aid}"))
                if items:
                    self._t_bulk_follow(scraper, items, 30, 90,
                                         self.var_dedup.get(), True)
            sess.submit_automation(task, (), f"⏰👤 定时关注「{kw}」")

    def on_show_scheduled(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("⏰ 定时任务管理")
        dlg.geometry("920x520")
        dlg.transient(self.root)

        ttk.Label(dlg, text="定时任务列表",
                  font=(FONT_UI, 13, "bold"),
                  foreground="#F25928").pack(anchor="w", padx=12, pady=(10, 2))
        ttk.Label(dlg, text="每天到指定时间自动执行，需保持软件运行",
                  foreground="#888").pack(anchor="w", padx=12, pady=(0, 6))

        cols = ("name", "account", "type", "schedule", "last_run", "enabled")
        tv = ttk.Treeview(dlg, columns=cols, show="headings", height=12)
        for c, label, w in [("name", "名称", 150), ("account", "账号", 90),
                            ("type", "类型", 140), ("schedule", "执行时间", 100),
                            ("last_run", "上次执行", 140), ("enabled", "状态", 70)]:
            tv.heading(c, text=label); tv.column(c, width=w, anchor="w")
        tv.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        def refresh():
            tv.delete(*tv.get_children())
            for t in sch.list_tasks():
                tv.insert("", "end", iid=t["id"], values=(
                    t["name"], t["account"],
                    sch.TASK_TYPES.get(t["type"], t["type"]),
                    t["schedule"], t.get("last_run", "") or "—",
                    "✓" if t.get("enabled") else "✗",
                ))
        refresh()

        bb = ttk.Frame(dlg); bb.pack(fill="x", padx=8, pady=4)

        def on_add():
            self._scheduled_edit(None, on_done=refresh)

        def on_edit():
            sel = tv.selection()
            if not sel: return
            tid = sel[0]
            t = next((x for x in sch.list_tasks() if x["id"] == tid), None)
            if t: self._scheduled_edit(t, on_done=refresh)

        def on_del():
            sel = tv.selection()
            if not sel: return
            if not messagebox.askyesno("确认", "删除该定时任务？"): return
            sch.delete_task(sel[0])
            refresh()

        def on_toggle():
            sel = tv.selection()
            if not sel: return
            t = next((x for x in sch.list_tasks() if x["id"] == sel[0]), None)
            if t:
                sch.update_task(t["id"], enabled=not t.get("enabled", True))
                refresh()

        ttk.Button(bb, text="+ 新建任务", command=on_add).pack(side="left", padx=4)
        ttk.Button(bb, text="✎ 编辑", command=on_edit).pack(side="left", padx=4)
        ttk.Button(bb, text="⏯ 启用/停用", command=on_toggle).pack(side="left", padx=4)
        ttk.Button(bb, text="🗑 删除", command=on_del).pack(side="left", padx=4)
        ttk.Button(bb, text="关闭", command=dlg.destroy).pack(side="right", padx=4)

    def _scheduled_edit(self, task=None, on_done=None):
        dlg = tk.Toplevel(self.root)
        dlg.title("编辑定时任务" if task else "新建定时任务")
        dlg.geometry("520x480")
        dlg.transient(self.root); dlg.grab_set()
        t = task or {}

        f = ttk.Frame(dlg, padding=14); f.pack(fill="both", expand=True)

        # ---- 基本信息 ----
        ttk.Label(f, text="任务名称:", font=(FONT_UI, 10, "bold")).grid(row=0, column=0, sticky="e", pady=6, padx=(0, 6))
        e_name = ttk.Entry(f, width=36); e_name.insert(0, t.get("name", "我的任务"))
        e_name.grid(row=0, column=1, columnspan=2, sticky="ew", pady=6)

        ttk.Label(f, text="执行账号:", font=(FONT_UI, 10, "bold")).grid(row=1, column=0, sticky="e", pady=6, padx=(0, 6))
        cb_acc = ttk.Combobox(f, values=list_accounts(), state="readonly", width=22)
        cb_acc.set(t.get("account", list_accounts()[0] if list_accounts() else ""))
        cb_acc.grid(row=1, column=1, sticky="w", pady=6)

        ttk.Label(f, text="任务类型:", font=(FONT_UI, 10, "bold")).grid(row=2, column=0, sticky="e", pady=6, padx=(0, 6))
        type_display = [v for v in sch.TASK_TYPES.values()]
        type_keys = list(sch.TASK_TYPES.keys())
        cb_type = ttk.Combobox(f, values=type_display, state="readonly", width=22)
        cur_type = t.get("type", "nurture")
        cb_type.set(sch.TASK_TYPES.get(cur_type, type_display[0]))
        cb_type.grid(row=2, column=1, sticky="w", pady=6)

        ttk.Label(f, text="每天执行时间:", font=(FONT_UI, 10, "bold")).grid(row=3, column=0, sticky="e", pady=6, padx=(0, 6))
        time_row = ttk.Frame(f); time_row.grid(row=3, column=1, sticky="w", pady=6)
        sched = t.get("schedule", "daily 09:00")
        time_str = sched.split(" ", 1)[1] if " " in sched else "09:00"
        e_time = ttk.Entry(time_row, width=8); e_time.insert(0, time_str)
        e_time.pack(side="left")
        ttk.Label(time_row, text="  (格式 HH:MM  如 09:00)", foreground="#888").pack(side="left")

        # ---- 参数区（根据类型动态切换） ----
        ttk.Separator(f, orient="horizontal").grid(row=4, column=0, columnspan=3, sticky="ew", pady=10)
        ttk.Label(f, text="⚙ 任务参数", font=(FONT_UI, 11, "bold"),
                  foreground="#F25928").grid(row=5, column=0, columnspan=3, sticky="w", pady=(0, 6))

        params_frame = ttk.Frame(f)
        params_frame.grid(row=6, column=0, columnspan=3, sticky="ew", pady=4)
        f.columnconfigure(1, weight=1)

        # 参数 widgets 存储
        param_widgets = {}
        old_params = t.get("params", {})

        def build_nurture_fields(pf):
            ttk.Label(pf, text="浏览时长 (分钟):").grid(row=0, column=0, sticky="w", pady=4, padx=(0, 8))
            e_dur = ttk.Entry(pf, width=10)
            e_dur.insert(0, str(old_params.get("duration", 30)))
            e_dur.grid(row=0, column=1, sticky="w", pady=4)
            ttk.Label(pf, text="建议 15~60", foreground="#888").grid(row=0, column=2, sticky="w", padx=8)

            ttk.Label(pf, text="点赞概率 (%):").grid(row=1, column=0, sticky="w", pady=4, padx=(0, 8))
            e_like = ttk.Entry(pf, width=10)
            e_like.insert(0, str(int(float(old_params.get("like_prob", 0.15)) * 100)))
            e_like.grid(row=1, column=1, sticky="w", pady=4)
            ttk.Label(pf, text="新号建议 5%，老号 15%", foreground="#888").grid(row=1, column=2, sticky="w", padx=8)

            ttk.Label(pf, text="收藏概率 (%):").grid(row=2, column=0, sticky="w", pady=4, padx=(0, 8))
            e_col = ttk.Entry(pf, width=10)
            e_col.insert(0, str(int(float(old_params.get("collect_prob", 0.05)) * 100)))
            e_col.grid(row=2, column=1, sticky="w", pady=4)
            ttk.Label(pf, text="新号建议 0%，老号 5%", foreground="#888").grid(row=2, column=2, sticky="w", padx=8)

            param_widgets["duration"] = e_dur
            param_widgets["like_prob"] = e_like
            param_widgets["collect_prob"] = e_col

        def build_search_like_fields(pf):
            ttk.Label(pf, text="搜索关键词:").grid(row=0, column=0, sticky="w", pady=4, padx=(0, 8))
            e_kw = ttk.Entry(pf, width=24)
            e_kw.insert(0, str(old_params.get("keyword", "")))
            e_kw.grid(row=0, column=1, sticky="w", pady=4)
            ttk.Label(pf, text="如: 减脂餐、穿搭", foreground="#888").grid(row=0, column=2, sticky="w", padx=8)

            ttk.Label(pf, text="搜索数量:").grid(row=1, column=0, sticky="w", pady=4, padx=(0, 8))
            e_cnt = ttk.Entry(pf, width=10)
            e_cnt.insert(0, str(old_params.get("count", 20)))
            e_cnt.grid(row=1, column=1, sticky="w", pady=4)

            ttk.Label(pf, text="最低点赞数:").grid(row=2, column=0, sticky="w", pady=4, padx=(0, 8))
            e_min = ttk.Entry(pf, width=10)
            e_min.insert(0, str(old_params.get("min_like", 1000)))
            e_min.grid(row=2, column=1, sticky="w", pady=4)
            ttk.Label(pf, text="过滤低互动笔记", foreground="#888").grid(row=2, column=2, sticky="w", padx=8)

            param_widgets["keyword"] = e_kw
            param_widgets["count"] = e_cnt
            param_widgets["min_like"] = e_min

        def build_follow_authors_fields(pf):
            ttk.Label(pf, text="搜索关键词:").grid(row=0, column=0, sticky="w", pady=4, padx=(0, 8))
            e_kw = ttk.Entry(pf, width=24)
            e_kw.insert(0, str(old_params.get("keyword", "")))
            e_kw.grid(row=0, column=1, sticky="w", pady=4)
            ttk.Label(pf, text="如: 美食、护肤", foreground="#888").grid(row=0, column=2, sticky="w", padx=8)

            ttk.Label(pf, text="关注数量:").grid(row=1, column=0, sticky="w", pady=4, padx=(0, 8))
            e_cnt = ttk.Entry(pf, width=10)
            e_cnt.insert(0, str(old_params.get("count", 10)))
            e_cnt.grid(row=1, column=1, sticky="w", pady=4)
            ttk.Label(pf, text="每次最多关注多少人", foreground="#888").grid(row=1, column=2, sticky="w", padx=8)

            param_widgets["keyword"] = e_kw
            param_widgets["count"] = e_cnt

        builders = {
            "nurture": build_nurture_fields,
            "search_like": build_search_like_fields,
            "follow_authors": build_follow_authors_fields,
        }

        def refresh_params(*_):
            nonlocal old_params
            # 切换类型时，如果是首次（新建）就重置参数；编辑时保留原值
            for w in params_frame.winfo_children():
                w.destroy()
            param_widgets.clear()
            idx = type_display.index(cb_type.get()) if cb_type.get() in type_display else 0
            tkey = type_keys[idx]
            builder = builders.get(tkey)
            if builder:
                builder(params_frame)

        cb_type.bind("<<ComboboxSelected>>", lambda e: _on_type_switch())

        def _on_type_switch():
            nonlocal old_params
            # 切换类型后重置参数为默认值
            idx = type_display.index(cb_type.get()) if cb_type.get() in type_display else 0
            tkey = type_keys[idx]
            defaults = {
                "nurture": {"duration": 30, "like_prob": 0.15, "collect_prob": 0.05},
                "search_like": {"keyword": "", "count": 20, "min_like": 1000},
                "follow_authors": {"keyword": "", "count": 10},
            }
            old_params = defaults.get(tkey, {})
            refresh_params()

        # 初始显示
        refresh_params()

        msg = tk.StringVar()
        ttk.Label(f, textvariable=msg, foreground="#c33",
                  font=(FONT_UI, 10)).grid(row=7, column=0, columnspan=3, sticky="w", pady=4)

        def collect_params():
            """从 UI 控件收集参数 dict"""
            idx = type_display.index(cb_type.get()) if cb_type.get() in type_display else 0
            tkey = type_keys[idx]
            p = {}
            if tkey == "nurture":
                p["duration"] = int(param_widgets["duration"].get() or 30)
                lp = int(param_widgets["like_prob"].get() or 15)
                cp = int(param_widgets["collect_prob"].get() or 5)
                p["like_prob"] = round(lp / 100, 2)
                p["collect_prob"] = round(cp / 100, 2)
            elif tkey == "search_like":
                p["keyword"] = param_widgets["keyword"].get().strip()
                p["count"] = int(param_widgets["count"].get() or 20)
                p["min_like"] = int(param_widgets["min_like"].get() or 1000)
                if not p["keyword"]:
                    raise ValueError("搜索关键词不能为空")
            elif tkey == "follow_authors":
                p["keyword"] = param_widgets["keyword"].get().strip()
                p["count"] = int(param_widgets["count"].get() or 10)
                if not p["keyword"]:
                    raise ValueError("搜索关键词不能为空")
            return p

        def do_save():
            name = e_name.get().strip()
            acc = cb_acc.get()
            type_idx = type_display.index(cb_type.get()) if cb_type.get() in type_display else 0
            ttype = type_keys[type_idx]
            time_s = e_time.get().strip()
            try:
                h, m = map(int, time_s.split(":"))
                assert 0 <= h < 24 and 0 <= m < 60
            except Exception:
                msg.set("时间格式应为 HH:MM，如 09:00"); return
            try:
                params = collect_params()
            except ValueError as e:
                msg.set(str(e)); return
            except Exception as e:
                msg.set(f"参数错误: {e}"); return
            if not name or not acc:
                msg.set("名称和账号必填"); return
            sched_str = f"daily {h:02d}:{m:02d}"
            if t.get("id"):
                sch.update_task(t["id"], name=name, account=acc, type=ttype,
                                params=params, schedule=sched_str)
            else:
                sch.add_task(name, acc, ttype, params, sched_str)
            dlg.destroy()
            if on_done: on_done()

        b = ttk.Frame(dlg); b.pack(pady=10)
        ttk.Button(b, text="✅ 保存", command=do_save).pack(side="left", padx=6)
        ttk.Button(b, text="取消", command=dlg.destroy).pack(side="left", padx=6)

    # ---------- 签名 JS 更新 ----------
    def _check_sign_update(self):
        try:
            updated, msg = sign_updater.check_and_update(
                log=lambda m: self.log("UPD", m))
            if updated:
                self.log("UPD", f"✓ {msg}")
            else:
                self.log("UPD", msg)
        except Exception as e:
            self.log("UPD", f"检查更新失败: {e}")

    def on_check_sign_update_manual(self):
        def run():
            self._check_sign_update()
            messagebox.showinfo("签名更新", "检查完成，详情看日志")
        threading.Thread(target=run, daemon=True).start()

    # ============ 软件版本更新 ============
    def _check_app_update_silent(self):
        """ 启动时静默检查；有更新弹提示，无更新不打扰 """
        try:
            info = app_updater.check_update()
            if info and info.get("has_update"):
                self.log("SYS", f"🆕 发现新版本 {info.get('latest_version')}，菜单 [📢 公告 → 🆕 检查软件更新]")
                # 强制升级 → 立刻弹窗，否则只打日志
                if info.get("force"):
                    self.root.after(1500, lambda: self._show_update_dialog(info))
        except Exception:
            pass

    def on_check_app_update(self):
        """ 用户主动点 [检查软件更新] """
        def run():
            try:
                info = app_updater.check_update()
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror(
                    "检查失败", f"无法连接服务器：{e}"))
                return
            if not info:
                self.root.after(0, lambda: messagebox.showerror(
                    "检查失败", "无法连接服务器"))
                return
            if not info.get("has_update"):
                self.root.after(0, lambda: messagebox.showinfo(
                    "已是最新版本",
                    f"当前版本：{app_updater.CURRENT_VERSION}\n"
                    f"服务器最新：{info.get('latest_version','未知')}\n"
                    "✓ 无需更新"))
                return
            self.root.after(0, lambda: self._show_update_dialog(info))
        threading.Thread(target=run, daemon=True).start()

    def _show_update_dialog(self, info):
        """ 显示版本升级对话框 """
        dlg = tk.Toplevel(self.root)
        dlg.title("🆕 发现新版本")
        dlg.geometry("580x520")
        dlg.minsize(560, 480)
        dlg.transient(self.root); dlg.grab_set()
        # 居中显示，确保按钮在屏幕内
        dlg.update_idletasks()
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        x = (sw - 580) // 2
        y = max(20, (sh - 520) // 2)
        dlg.geometry(f"580x520+{x}+{y}")

        force = bool(info.get("force"))
        title_text = "🚨 强制更新" if force else "🆕 发现新版本"
        ttk.Label(dlg, text=title_text,
                  font=(FONT_UI, 14, "bold"),
                  foreground=("#c33" if force else "#F25928")).pack(pady=(12, 4))

        ttk.Label(dlg, text=f"当前版本：{app_updater.CURRENT_VERSION}    "
                            f"最新版本：{info.get('latest_version','')}",
                  font=(FONT_UI, 11)).pack(pady=4)

        # ⚠ 先把底部区域（按钮+进度条+大小）pack 到底部，再 pack 中间的滚动文本
        #    这样按钮永远可见，不会被 Text expand 挤出去

        # 强制更新提示（如果有）放底部
        force_hint_frame = ttk.Frame(dlg)
        force_hint_frame.pack(side="bottom", fill="x", pady=(0, 6))
        if force:
            ttk.Label(force_hint_frame,
                      text="⚠ 此为强制更新，必须升级后才能继续使用",
                      foreground="#c33").pack()

        # 按钮区
        bb = ttk.Frame(dlg)
        bb.pack(side="bottom", pady=10)

        # 状态消息
        msg_var = tk.StringVar()
        ttk.Label(dlg, textvariable=msg_var, foreground="#0a7").pack(side="bottom")

        # 进度条
        prog_var = tk.IntVar(value=0)
        prog = ttk.Progressbar(dlg, variable=prog_var, maximum=100, length=520)
        prog.pack(side="bottom", padx=20, pady=(8, 4))

        # 文件大小
        size_mb = (info.get("size") or 0) / 1024 / 1024
        ttk.Label(dlg, text=f"大小：{size_mb:.1f} MB",
                  foreground="#888").pack(side="bottom", anchor="w", padx=20)

        # 中间是滚动文本（更新内容）
        ttk.Label(dlg, text="更新内容：",
                  font=(FONT_UI, 11, "bold")).pack(anchor="w", padx=20, pady=(8, 2))
        cl = tk.Text(dlg, height=8, font=(FONT_UI, 10),
                     bg="#f7f9fc", relief="flat", wrap="word")
        cl.pack(fill="both", expand=True, padx=20, pady=4)
        cl.insert("1.0", info.get("changelog") or "(无更新说明)")
        cl.config(state="disabled")

        def start_update():
            btn_up.configure(state="disabled")
            btn_later.configure(state="disabled")
            msg_var.set("下载中...")
            def run():
                try:
                    def prog_cb(done, total):
                        if total > 0:
                            pct = int(done * 100 / total)
                            self.root.after(0, lambda: prog_var.set(pct))
                            self.root.after(0, lambda: msg_var.set(
                                f"下载 {done/1024/1024:.1f} / {total/1024/1024:.1f} MB"))
                    new_exe = app_updater.download_update(info, progress_cb=prog_cb)
                    self.root.after(0, lambda: msg_var.set("下载完成，准备替换..."))
                    time.sleep(0.5)
                    self.root.after(0, lambda: msg_var.set(
                        "✓ 已下载，3 秒后自动重启..."))
                    self.root.after(2500, lambda: app_updater.apply_update(new_exe))
                except RuntimeError as e:
                    self.root.after(0, lambda: msg_var.set(f"✗ {e}"))
                    self.root.after(0, lambda: btn_up.configure(state="normal"))
                    self.root.after(0, lambda: btn_later.configure(state="normal"))
                except Exception as e:
                    self.root.after(0, lambda: msg_var.set(f"✗ 更新失败: {e}"))
                    self.root.after(0, lambda: btn_up.configure(state="normal"))
                    self.root.after(0, lambda: btn_later.configure(state="normal"))
            threading.Thread(target=run, daemon=True).start()

        btn_up = ttk.Button(bb, text="🚀 立即更新", command=start_update)
        btn_up.pack(side="left", padx=6)
        btn_later = ttk.Button(bb, text="稍后再说",
                                command=dlg.destroy if not force else None,
                                state=("disabled" if force else "normal"))
        btn_later.pack(side="left", padx=6)

    # ---------- 公告 ----------
    def _fetch_announces_on_start(self):
        try:
            news = announce_client.fetch_new()
        except Exception:
            news = []
        if news:
            # 主线程弹窗
            self.root.after(500, lambda: self._show_announce_popup(news))

    def _show_announce_popup(self, announces):
        if not announces:
            return
        # 倒序的，最旧的在最后 → 先标记 read 最大的
        try:
            announce_client.mark_read(max(a["id"] for a in announces))
        except Exception:
            pass
        dlg = tk.Toplevel(self.root)
        dlg.title(f"📢 公告 ({len(announces)} 条新消息)")
        dlg.geometry("520x420")
        dlg.transient(self.root); dlg.lift()
        dlg.attributes("-topmost", True)
        dlg.after(800, lambda: dlg.attributes("-topmost", False))
        body = tk.Text(dlg, bg="#1e1e1e", fg="#d4d4d4",
                       font=(FONT_UI, 10), wrap="word", padx=14, pady=10)
        body.pack(fill="both", expand=True)
        for a in announces:
            prio = {"info": "📝 ", "warn": "⚠ ", "urgent": "🚨 "}.get(a.get("priority"), "")
            body.insert("end", f"{prio}{a.get('title','')}\n", "title")
            body.insert("end", f"{a.get('content','')}\n\n", "")
            body.insert("end", "─" * 50 + "\n\n", "sep")
        body.tag_config("title", foreground="#7ee787",
                        font=(FONT_UI, 11, "bold"))
        body.tag_config("sep", foreground="#444")
        body.config(state="disabled")
        ttk.Button(dlg, text="知道了", command=dlg.destroy).pack(pady=8)

    def on_show_announces(self):
        threading.Thread(
            target=lambda: self._show_announce_popup(announce_client.fetch_new()),
            daemon=True,
        ).start()

    # ---------- 备份/恢复 ----------
    def on_backup(self):
        import zipfile
        from tkinter import filedialog
        f = filedialog.asksaveasfilename(
            defaultextension=".zip",
            initialfile=f"qiqi_backup_{time.strftime('%Y%m%d_%H%M%S')}.zip",
            filetypes=[("ZIP 压缩包", "*.zip")],
        )
        if not f:
            return
        base = Path(os.path.dirname(os.path.abspath(__file__)))
        try:
            with zipfile.ZipFile(f, "w", zipfile.ZIP_DEFLATED) as zf:
                # accounts/ 全部
                acc_dir = base / "accounts"
                if acc_dir.exists():
                    for p in acc_dir.glob("*"):
                        if p.is_file():
                            zf.write(p, f"accounts/{p.name}")
                # license_server.txt（指向你服务器）
                lf = base / "license_server.txt"
                if lf.exists():
                    zf.write(lf, "license_server.txt")
                # 注意：~/.qiqi_license.dat 不再放入备份。激活绑定当前机器，
                # 跨机搬运没有意义且会触发"机器标识变更"错误。
            self.log("SYS", f"✓ 备份完成: {f}")
            messagebox.showinfo("备份完成", f"已导出到:\n{f}")
        except Exception as e:
            messagebox.showerror("备份失败", str(e))

    def on_restore(self):
        import zipfile
        from tkinter import filedialog
        f = filedialog.askopenfilename(
            filetypes=[("ZIP 压缩包", "*.zip")],
        )
        if not f:
            return
        if not messagebox.askyesno(
            "确认恢复",
            "将覆盖现有的:\n"
            "  - accounts/ 目录（所有账号 cookie/代理配置）\n"
            "  - license_server.txt\n\n"
            "授权信息绑定当前机器，不会从备份恢复。\n"
            "请先关闭所有账号浏览器（避免冲突）。确认继续？"
        ):
            return
        # 关闭所有 session
        for s in list(self.sessions.values()):
            try: s.shutdown()
            except Exception: pass
        self.sessions.clear()

        base = Path(os.path.dirname(os.path.abspath(__file__)))
        try:
            with zipfile.ZipFile(f, "r") as zf:
                for name in zf.namelist():
                    if name.endswith("/"):
                        continue
                    # 旧版本 ZIP 里可能含有 _user/qiqi_license.dat — 忽略，
                    # 防止把别人机器的授权文件搬过来导致激活漂移
                    if name.startswith("_user/"):
                        continue
                    if name.startswith("accounts/") or name == "license_server.txt":
                        # 防穿越
                        if ".." in name:
                            continue
                        out = base / name
                        out.parent.mkdir(parents=True, exist_ok=True)
                        out.write_bytes(zf.read(name))
            # 清缓存
            from scraper import load_config
            import settings_mgr
            settings_mgr._cache = None
            self.log("SYS", f"✓ 从 {f} 恢复完成")
            messagebox.showinfo(
                "恢复完成",
                "数据已恢复。建议重启程序使设置完全生效。",
            )
            self._refresh_accounts()
        except Exception as e:
            messagebox.showerror("恢复失败", str(e))

    def on_show_captured(self):
        """ 显示当前选中账号最近抓到的写操作 POST 请求 """
        name = self.acc_var.get()
        sess = self.sessions.get(name)
        if not sess or not sess.scraper:
            return messagebox.showinfo("提示", "请先启动该账号的浏览器")
        reqs = getattr(sess.scraper, "captured_requests", [])
        if not reqs:
            return messagebox.showinfo(
                "提示",
                "暂无抓到的请求。\n\n请在该账号的浏览器窗口里：\n"
                "1. 打开任意一条小红书笔记\n"
                "2. 手动评论一条（或点赞/关注）\n"
                "3. 回来再点这个菜单看抓到了什么"
            )
        dlg = tk.Toplevel(self.root)
        dlg.title(f"📡 抓到的 API 请求 - {name}  共 {len(reqs)} 条")
        dlg.geometry("980x680")
        dlg.transient(self.root)

        lst_frame = ttk.Frame(dlg); lst_frame.pack(side="left", fill="y", padx=4, pady=4)
        lb = tk.Listbox(lst_frame, width=46, height=30, font=(FONT_MONO, 9))
        lb.pack(side="left", fill="y")
        sb = ttk.Scrollbar(lst_frame, command=lb.yview); sb.pack(side="left", fill="y")
        lb.config(yscrollcommand=sb.set)
        for i, r in enumerate(reqs):
            short = r["url"].split("?")[0].split("/api/sns/web/")[-1][:38]
            lb.insert("end", f"[{r['ts']}] {r['method']} {short}")

        txt = tk.Text(dlg, font=(FONT_MONO, 9), bg="#1e1e1e", fg="#d4d4d4", wrap="word")
        txt.pack(side="left", fill="both", expand=True, padx=4, pady=4)

        def show(_evt=None):
            sel = lb.curselection()
            if not sel: return
            r = reqs[sel[0]]
            txt.delete("1.0", "end")
            txt.insert("end", f"URL:\n  {r['url']}\n\n")
            txt.insert("end", f"Method: {r['method']}\n\n")
            txt.insert("end", "Key Headers:\n")
            for k, v in r["key_headers"].items():
                short_v = v[:200] + ("..." if len(v) > 200 else "")
                txt.insert("end", f"  {k}: {short_v}\n")
            txt.insert("end", "\n所有 Header keys:\n  " + ", ".join(r["all_header_keys"]) + "\n\n")
            txt.insert("end", f"Body:\n{r['body']}\n")
        lb.bind("<<ListboxSelect>>", show)
        if reqs:
            lb.select_set(len(reqs) - 1); show()

        def copy_all():
            r = reqs[lb.curselection()[0]] if lb.curselection() else reqs[-1]
            import json
            text = json.dumps(r, ensure_ascii=False, indent=2)
            self.root.clipboard_clear(); self.root.clipboard_append(text)
            self.log("SYS", "已复制选中请求 JSON 到剪贴板")
        ttk.Button(dlg, text="复制当前请求 JSON", command=copy_all).pack(side="bottom", pady=4)

    def on_show_license(self):
        d = self.license_data or {}
        dlg = tk.Toplevel(self.root)
        dlg.title("🔑 授权信息")
        dlg.geometry("420x280")
        dlg.transient(self.root)
        f = ttk.Frame(dlg, padding=18); f.pack(fill="both", expand=True)
        rows = [
            ("套餐", lm.PLAN_NAMES.get(d.get("plan", ""), d.get("plan", "—"))),
            ("卡密", d.get("key", "")),
            ("机器码", d.get("machine_id", "")),
            ("到期时间", lm.fmt_expire(d["expires_at"]) if d.get("expires_at") else "—"),
            ("剩余时长", lm.fmt_remain(d["expires_at"]) if d.get("expires_at") else "—"),
        ]
        for i, (k, v) in enumerate(rows):
            ttk.Label(f, text=k + ":", foreground="#888").grid(row=i, column=0, sticky="ne", padx=4, pady=4)
            tk.Entry(f, font=(FONT_MONO, 9), bd=0, readonlybackground="#fff",
                     state="readonly", relief="flat", width=42).grid(row=i, column=1, sticky="ew", pady=4)
            f.grid_slaves(row=i, column=1)[0].configure(state="normal")
            f.grid_slaves(row=i, column=1)[0].insert(0, str(v))
            f.grid_slaves(row=i, column=1)[0].configure(state="readonly")
        f.columnconfigure(1, weight=1)
        ttk.Button(dlg, text="关闭", command=dlg.destroy).pack(pady=8)

    def on_show_stats(self):
        if not db.is_enabled():
            return messagebox.showinfo("提示", "未配置 MySQL，无法统计")
        dlg = tk.Toplevel(self.root)
        dlg.title("📊 今日操作统计")
        dlg.geometry("440x300")
        dlg.transient(self.root)
        tv = ttk.Treeview(dlg, columns=("acc", "like", "comment", "follow", "total"),
                          show="headings", height=12)
        for c, label, w in [("acc", "账号", 100), ("like", "点赞", 60),
                            ("comment", "评论", 60), ("follow", "关注", 60),
                            ("total", "合计", 60)]:
            tv.heading(c, text=label); tv.column(c, width=w, anchor="center")
        tv.pack(fill="both", expand=True, padx=8, pady=8)

        total = {"like": 0, "comment": 0, "follow": 0}
        for a in list_accounts():
            s = db.today_stats(a)
            row = (a, s.get("like", 0), s.get("comment", 0), s.get("follow", 0),
                   sum(s.values()))
            tv.insert("", "end", values=row)
            for k in total:
                total[k] += s.get(k, 0)
        tv.insert("", "end", values=("合计", total["like"], total["comment"],
                                       total["follow"], sum(total.values())),
                  tags=("total",))
        tv.tag_configure("total", background="#2c3e50", foreground="#fff")
        ttk.Button(dlg, text="关闭", command=dlg.destroy).pack(pady=4)

    def _build_tabs(self, nb):
        # ① 搜索
        f1 = ttk.Frame(nb, padding=8); nb.add(f1, text="① 搜索")
        ttk.Label(f1, text="关键词:").grid(row=0, column=0, sticky="w")
        self.e_kw = ttk.Entry(f1, width=30); self.e_kw.grid(row=0, column=1, padx=4)
        ttk.Label(f1, text="数量:").grid(row=0, column=2)
        self.e_count = ttk.Entry(f1, width=8); self.e_count.insert(0, "20"); self.e_count.grid(row=0, column=3, padx=4)
        ttk.Button(f1, text="🔍 开始搜索", command=self.on_search).grid(row=0, column=4, padx=8)

        # 搜索时自动采 IP 属地
        self.var_search_with_ip = tk.BooleanVar(value=False)
        ttk.Checkbutton(f1, text="🌍 顺便采 IP 属地（每篇 +1.5s，结果立刻填充）",
                        variable=self.var_search_with_ip).grid(row=1, column=0, columnspan=5,
                                                                sticky="w", pady=4)
        ttk.Label(f1, text="结果会显示在下方面板，可勾选批量操作",
                  foreground="#888").grid(row=2, column=0, columnspan=6, sticky="w", pady=4)

        # ② 笔记+评论
        f2 = ttk.Frame(nb, padding=8); nb.add(f2, text="② 笔记+评论")
        ttk.Label(f2, text="笔记 URL/ID:").grid(row=0, column=0, sticky="w")
        self.e_note = ttk.Entry(f2, width=55); self.e_note.grid(row=0, column=1, columnspan=4, padx=4)
        ttk.Label(f2, text="评论上限:").grid(row=1, column=0, sticky="w", pady=4)
        self.e_cmt = ttk.Entry(f2, width=8); self.e_cmt.insert(0, "100"); self.e_cmt.grid(row=1, column=1, sticky="w", pady=4)
        self.var_dl_media = tk.BooleanVar(value=False)
        ttk.Checkbutton(f2, text="📥 下载图片/视频", variable=self.var_dl_media).grid(row=1, column=2, sticky="w")
        self.var_extract_intent = tk.BooleanVar(value=True)
        ttk.Checkbutton(f2, text="🎯 提取意向用户", variable=self.var_extract_intent).grid(row=1, column=3, sticky="w")
        ttk.Button(f2, text="开始采集", command=self.on_note).grid(row=1, column=4, sticky="w", padx=8)

        # 自动回复意向评论
        self.var_auto_reply = tk.BooleanVar(value=False)
        ttk.Checkbutton(f2, text="✨ 自动回复意向评论（求/怎么买/求链接 等潜客 - 需 API 模式）",
                        variable=self.var_auto_reply).grid(row=2, column=0, columnspan=5, sticky="w", pady=(8, 2))
        ttk.Label(f2, text="意向回复模板（每行一条；启用 AI 时仅作为风格参考，AI 会基于潜客原话生成）:",
                  foreground="#888").grid(row=3, column=0, columnspan=5, sticky="w", pady=(4, 0))
        self.txt_reply = tk.Text(f2, height=4, width=100, font=(FONT_UI, 11))
        self.txt_reply.grid(row=4, column=0, columnspan=5, padx=2, pady=4)
        self.txt_reply.insert("1.0",
            "私我哈~\n姐妹 dd 我聊\n主页有详情可以看下\n私聊我哦\n扣 1 我私\n这边私~\n看简介找我")
        ttk.Label(f2, text="💡 本 tab 所有设置同样作用于 [📥 批量采笔记+评论] 按钮 - 直接在结果面板勾选多条批量执行",
                  foreground="#0a7").grid(row=5, column=0, columnspan=5, sticky="w", pady=(6, 2))

        # ③ 用户主页
        f3 = ttk.Frame(nb, padding=8); nb.add(f3, text="③ 用户主页")
        ttk.Label(f3, text="用户 URL/ID:").grid(row=0, column=0, sticky="w")
        self.e_user = ttk.Entry(f3, width=55); self.e_user.grid(row=0, column=1, columnspan=3, padx=4)
        ttk.Label(f3, text="笔记数:").grid(row=1, column=0, sticky="w", pady=4)
        self.e_ucount = ttk.Entry(f3, width=8); self.e_ucount.insert(0, "30"); self.e_ucount.grid(row=1, column=1, sticky="w", pady=4)
        ttk.Button(f3, text="开始采集", command=self.on_user).grid(row=1, column=2, sticky="w", padx=8)

        # ④ 评论助手
        f4 = ttk.Frame(nb, padding=8); nb.add(f4, text="④ 评论模板 ⚠")
        ttk.Label(f4, text="评论模板（每行一条 随机抽 建议≥10条）:").grid(row=0, column=0, columnspan=6, sticky="w")
        self.txt_templates = tk.Text(f4, height=5, width=100, font=(FONT_UI, 11))
        self.txt_templates.grid(row=1, column=0, columnspan=6, padx=2, pady=4)
        self.txt_templates.insert("1.0",
            "学到了 谢谢分享~\n好喜欢这种风格啊\n请问博主用的是什么呢\n码住码住 周末试试\n"
            "怎么这么好看 求链接\n刚关注 期待更多内容\n看完整个人都被治愈了\n终于刷到干货了\n"
            "宝藏博主 已三连\n这个视角好新颖 收藏了\n")
        o = ttk.Frame(f4); o.grid(row=2, column=0, columnspan=6, sticky="w", pady=4)
        ttk.Label(o, text="间隔 最小:").pack(side="left")
        self.e_dmin = ttk.Entry(o, width=5); self.e_dmin.insert(0, "60"); self.e_dmin.pack(side="left")
        ttk.Label(o, text="最大:").pack(side="left", padx=(6, 0))
        self.e_dmax = ttk.Entry(o, width=5); self.e_dmax.insert(0, "150"); self.e_dmax.pack(side="left")
        ttk.Label(o, text="    今日上限:").pack(side="left", padx=(12, 0))
        self.e_daily = ttk.Entry(o, width=5); self.e_daily.insert(0, "20"); self.e_daily.pack(side="left")
        self.var_confirm = tk.BooleanVar(value=False)
        ttk.Checkbutton(o, text="每条手动确认", variable=self.var_confirm).pack(side="left", padx=12)
        self.var_shuffle = tk.BooleanVar(value=True)
        ttk.Checkbutton(o, text="打乱顺序", variable=self.var_shuffle).pack(side="left")
        self.var_use_ai = tk.BooleanVar(value=False)
        ttk.Checkbutton(o, text="✨ AI 改写每条", variable=self.var_use_ai).pack(side="left", padx=12)
        self.var_dedup = tk.BooleanVar(value=True)
        ttk.Checkbutton(o, text="🗄 自动去重（需 MySQL）", variable=self.var_dedup).pack(side="left")
        ttk.Label(f4, text="↑ 模板设好后，结果面板勾选目标 → 点 [批量评论]。AI 改写会用 DeepSeek 给每条评论生成个性化文案。",
                  foreground="#0a7").grid(row=3, column=0, columnspan=6, sticky="w", pady=4)

        # ⑤ 互动参数（关注/点赞）
        f5 = ttk.Frame(nb, padding=8); nb.add(f5, text="⑤ 互动参数 ⚠")
        ttk.Label(f5, text="关注/点赞 间隔 最小:").grid(row=0, column=0, sticky="w")
        self.e_fmin = ttk.Entry(f5, width=6); self.e_fmin.insert(0, "30"); self.e_fmin.grid(row=0, column=1)
        ttk.Label(f5, text="最大:").grid(row=0, column=2)
        self.e_fmax = ttk.Entry(f5, width=6); self.e_fmax.insert(0, "90"); self.e_fmax.grid(row=0, column=3)
        ttk.Label(f5, text="单次上限:").grid(row=0, column=4)
        self.e_flim = ttk.Entry(f5, width=6); self.e_flim.insert(0, "30"); self.e_flim.grid(row=0, column=5)
        ttk.Label(f5, text="↑ 设置好后，在结果面板勾选 → 点 [勾选 → 批量点赞 / 关注作者]",
                  foreground="#0a7").grid(row=1, column=0, columnspan=6, sticky="w", pady=8)

        # ⑥ 爆品发现
        f6 = ttk.Frame(nb, padding=8); nb.add(f6, text="⑥ 爆品发现 🔥")
        ttk.Label(f6, text="关键词:").grid(row=0, column=0, sticky="w")
        self.e_hkw = ttk.Entry(f6, width=22); self.e_hkw.grid(row=0, column=1, padx=4)
        ttk.Label(f6, text="扫描量:").grid(row=0, column=2)
        self.e_hscan = ttk.Entry(f6, width=6); self.e_hscan.insert(0, "100"); self.e_hscan.grid(row=0, column=3)
        ttk.Label(f6, text="取前N:").grid(row=0, column=4)
        self.e_htop = ttk.Entry(f6, width=6); self.e_htop.insert(0, "30"); self.e_htop.grid(row=0, column=5)
        ttk.Button(f6, text="🔥 挖爆品", command=self.on_hot).grid(row=0, column=6, padx=8)

        ttk.Label(f6, text="最低点赞:").grid(row=1, column=0, sticky="w", pady=4)
        self.e_hmin_like = ttk.Entry(f6, width=10); self.e_hmin_like.insert(0, "1000"); self.e_hmin_like.grid(row=1, column=1, sticky="w")
        ttk.Label(f6, text="最低评论:").grid(row=1, column=2)
        self.e_hmin_cmt = ttk.Entry(f6, width=10); self.e_hmin_cmt.insert(0, "0"); self.e_hmin_cmt.grid(row=1, column=3)
        ttk.Label(f6, text="最低收藏:").grid(row=1, column=4)
        self.e_hmin_col = ttk.Entry(f6, width=10); self.e_hmin_col.insert(0, "0"); self.e_hmin_col.grid(row=1, column=5)
        ttk.Label(f6, text="排序:").grid(row=2, column=0, sticky="w", pady=4)
        self.cb_hsort = ttk.Combobox(f6, values=["按点赞", "按评论", "按收藏"], state="readonly", width=10)
        self.cb_hsort.set("按点赞"); self.cb_hsort.grid(row=2, column=1, sticky="w")

        # ⑧ 笔记发布
        f8 = ttk.Frame(nb, padding=8); nb.add(f8, text="⑧ 发布笔记 📝")

        # 类型 + AI 生成
        type_row = ttk.Frame(f8); type_row.grid(row=0, column=0, columnspan=4, sticky="w", pady=4)
        ttk.Label(type_row, text="类型:").pack(side="left")
        self.var_pub_type = tk.StringVar(value="image")
        for v, label in [("image", "📷 图文"), ("video", "🎬 视频"), ("longtext", "📰 长文")]:
            ttk.Radiobutton(type_row, text=label, value=v,
                            variable=self.var_pub_type,
                            command=self._on_pub_type_change).pack(side="left", padx=4)
        ttk.Label(type_row, text="    ✨ AI 生成主题:").pack(side="left", padx=(20, 4))
        self.e_pub_topic = ttk.Entry(type_row, width=18)
        self.e_pub_topic.pack(side="left", padx=2)
        ttk.Label(type_row, text="风格:").pack(side="left")
        self.e_pub_style = ttk.Entry(type_row, width=14)
        self.e_pub_style.pack(side="left", padx=2)
        ttk.Button(type_row, text="✨ AI 生成内容", command=self.on_pub_ai_gen).pack(side="left", padx=6)

        ttk.Label(f8, text="标题:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
        self.e_pub_title = ttk.Entry(f8, width=70)
        self.e_pub_title.grid(row=1, column=1, columnspan=3, sticky="ew", pady=4)
        ttk.Label(f8, text="正文:").grid(row=2, column=0, sticky="ne", padx=4, pady=4)
        self.txt_pub_body = tk.Text(f8, height=6, width=80, font=(FONT_UI, 11))
        self.txt_pub_body.grid(row=2, column=1, columnspan=3, sticky="ew", pady=4)
        ttk.Label(f8, text="话题标签:").grid(row=3, column=0, sticky="e", padx=4)
        self.e_pub_tags = ttk.Entry(f8, width=70)
        self.e_pub_tags.grid(row=3, column=1, columnspan=3, sticky="ew", pady=4)
        ttk.Label(f8, text="多个用逗号: 早餐,减脂餐,健康",
                  foreground="#888").grid(row=4, column=1, sticky="w")

        # 媒体 (图片/视频)
        self.lbl_pub_media = ttk.Label(f8, text="图片:")
        self.lbl_pub_media.grid(row=5, column=0, sticky="ne", padx=4, pady=4)
        pic_frame = ttk.Frame(f8); pic_frame.grid(row=5, column=1, columnspan=3, sticky="ew", pady=4)
        self.lst_pub_imgs = tk.Listbox(pic_frame, height=4, width=80)
        self.lst_pub_imgs.pack(side="left", fill="x", expand=True)
        pic_btns = ttk.Frame(pic_frame); pic_btns.pack(side="left", padx=4)
        ttk.Button(pic_btns, text="+ 添加", command=self._pub_add_imgs).pack(fill="x", pady=1)
        ttk.Button(pic_btns, text="清空", command=self._pub_clear_imgs).pack(fill="x", pady=1)
        self.pic_frame_ref = pic_frame

        # 多账号选择
        ttk.Label(f8, text="发布到账号:").grid(row=6, column=0, sticky="ne", padx=4, pady=8)
        acc_frame = ttk.Frame(f8); acc_frame.grid(row=6, column=1, columnspan=3, sticky="w", pady=8)
        self._pub_acc_vars = {}
        self._pub_acc_frame = acc_frame

        bf = ttk.Frame(f8); bf.grid(row=7, column=1, sticky="w", pady=8)
        ttk.Button(bf, text="🔄 刷新账号列表", command=self._refresh_pub_accounts).pack(side="left", padx=2)
        ttk.Button(bf, text="📝 发布到选中账号", command=self.on_publish).pack(side="left", padx=8)
        ttk.Label(f8, text="间隔 30-60s 随机（多号防关联）",
                  foreground="#888").grid(row=8, column=1, sticky="w")
        self._refresh_pub_accounts()
        f8.columnconfigure(1, weight=1)

        # ⑦ 自动养号
        f7 = ttk.Frame(nb, padding=8); nb.add(f7, text="⑦ 自动养号 🌱")
        ttk.Label(f7, text="模拟真人浏览：开 explore → 随机进笔记 → 滚正文+评论 → 偶尔点赞收藏 → 退回",
                  foreground="#0a7").grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 6))
        ttk.Label(f7, text="时长(分钟):").grid(row=1, column=0, sticky="w")
        self.e_nur_dur = ttk.Entry(f7, width=8); self.e_nur_dur.insert(0, "30")
        self.e_nur_dur.grid(row=1, column=1, padx=4)
        ttk.Label(f7, text="点赞概率(%):").grid(row=1, column=2)
        self.e_nur_like = ttk.Entry(f7, width=8); self.e_nur_like.insert(0, "15")
        self.e_nur_like.grid(row=1, column=3, padx=4)
        ttk.Label(f7, text="收藏概率(%):").grid(row=1, column=4)
        self.e_nur_col = ttk.Entry(f7, width=8); self.e_nur_col.insert(0, "5")
        self.e_nur_col.grid(row=1, column=5, padx=4)
        ttk.Button(f7, text="🌱 开始养号", command=self.on_nurture).grid(row=2, column=0, columnspan=2, sticky="w", pady=8)
        ttk.Label(f7,
                  text="新号(<7天)建议: 时长60min、点赞5%、不收藏  |  "
                       "老号孵化: 30min、15%、5%",
                  foreground="#888").grid(row=2, column=2, columnspan=4, sticky="w")

        # ⑨ 自动搬运 🤖
        f9 = ttk.Frame(nb, padding=8); nb.add(f9, text="⑨ 自动搬运 🤖")
        ttk.Label(f9,
                  text="搜索关键词 → 下载素材 → 图片去重处理 → DeepSeek 改写文案 → 半自动/全自动发布",
                  foreground="#0a7",
                  font=(FONT_UI, 10, "bold")).grid(row=0, column=0, columnspan=8, sticky="w", pady=(0, 4))
        ttk.Label(f9,
                  text="⚠ 法律提示：搬运他人作品涉嫌侵犯著作权/肖像权，请仅用于自有素材或灵感参考。"
                       "本功能默认半自动（生成完先预览，由你确认发布）。",
                  foreground="#c33").grid(row=1, column=0, columnspan=8, sticky="w", pady=(0, 6))

        # 第一行：关键词 + 类型 + 数量
        ttk.Label(f9, text="搜索关键词:").grid(row=2, column=0, sticky="e")
        self.e_rp_kw = ttk.Entry(f9, width=18); self.e_rp_kw.insert(0, "风景")
        self.e_rp_kw.grid(row=2, column=1, padx=4, sticky="w")
        ttk.Label(f9, text="素材类型:").grid(row=2, column=2, sticky="e")
        self.cb_rp_type = ttk.Combobox(f9, values=["风景", "人物", "动物", "美食", "穿搭", "其他"],
                                        state="readonly", width=8)
        self.cb_rp_type.set("风景"); self.cb_rp_type.grid(row=2, column=3, padx=4, sticky="w")
        ttk.Label(f9, text="搬运数量:").grid(row=2, column=4, sticky="e")
        self.e_rp_count = ttk.Entry(f9, width=6); self.e_rp_count.insert(0, "3")
        self.e_rp_count.grid(row=2, column=5, padx=4, sticky="w")
        ttk.Label(f9, text="(建议单号≤3/日)", foreground="#888").grid(row=2, column=6, columnspan=2, sticky="w")

        # 第二行：最低点赞 + 风格描述
        ttk.Label(f9, text="原图最低点赞:").grid(row=3, column=0, sticky="e")
        self.e_rp_min_like = ttk.Entry(f9, width=10); self.e_rp_min_like.insert(0, "1000")
        self.e_rp_min_like.grid(row=3, column=1, padx=4, sticky="w")
        ttk.Label(f9, text="文案风格:").grid(row=3, column=2, sticky="e")
        self.e_rp_style = ttk.Entry(f9, width=24)
        self.e_rp_style.insert(0, "")
        self.e_rp_style.grid(row=3, column=3, columnspan=3, padx=4, sticky="ew")
        ttk.Label(f9, text="(留空按类型自动选)", foreground="#888").grid(row=3, column=6, columnspan=2, sticky="w")

        # 第三行：图片处理开关
        proc_row = ttk.Frame(f9); proc_row.grid(row=4, column=0, columnspan=8, sticky="w", pady=(8, 2))
        ttk.Label(proc_row, text="🖼 图片处理（必做，否则平台秒识别）:",
                  font=(FONT_UI, 10, "bold")).pack(side="left")
        self.var_rp_noise = tk.BooleanVar(value=True)
        ttk.Checkbutton(proc_row, text="加噪点", variable=self.var_rp_noise).pack(side="left", padx=6)
        self.var_rp_crop = tk.BooleanVar(value=True)
        ttk.Checkbutton(proc_row, text="微裁边", variable=self.var_rp_crop).pack(side="left", padx=6)
        self.var_rp_wm = tk.BooleanVar(value=False)
        ttk.Checkbutton(proc_row, text="加水印", variable=self.var_rp_wm).pack(side="left", padx=6)
        ttk.Label(proc_row, text="水印文字:").pack(side="left", padx=(8, 2))
        self.e_rp_wm_text = ttk.Entry(proc_row, width=12); self.e_rp_wm_text.insert(0, "")
        self.e_rp_wm_text.pack(side="left")
        self.var_rp_rand_name = tk.BooleanVar(value=True)
        ttk.Checkbutton(proc_row, text="随机文件名",
                        variable=self.var_rp_rand_name).pack(side="left", padx=6)

        # 第四行：发布模式 + 发布间隔
        mode_row = ttk.Frame(f9); mode_row.grid(row=5, column=0, columnspan=8, sticky="w", pady=(4, 2))
        ttk.Label(mode_row, text="🚀 发布模式:",
                  font=(FONT_UI, 10, "bold")).pack(side="left")
        self.var_rp_mode = tk.StringVar(value="semi")
        ttk.Radiobutton(mode_row, text="半自动（预览审核后发布）",
                        value="semi", variable=self.var_rp_mode).pack(side="left", padx=6)
        ttk.Radiobutton(mode_row, text="全自动（高风险，自动发布全部）",
                        value="auto", variable=self.var_rp_mode).pack(side="left", padx=6)
        ttk.Label(mode_row, text="  发布间隔:").pack(side="left", padx=(16, 2))
        self.e_rp_dmin = ttk.Entry(mode_row, width=5); self.e_rp_dmin.insert(0, "300")
        self.e_rp_dmin.pack(side="left")
        ttk.Label(mode_row, text="~").pack(side="left")
        self.e_rp_dmax = ttk.Entry(mode_row, width=5); self.e_rp_dmax.insert(0, "900")
        self.e_rp_dmax.pack(side="left")
        ttk.Label(mode_row, text="秒（仅全自动）").pack(side="left", padx=2)

        # 执行按钮
        btn_row = ttk.Frame(f9); btn_row.grid(row=6, column=0, columnspan=8, sticky="w", pady=10)
        ttk.Button(btn_row, text="🤖 开始搬运（生成预览）",
                   command=self.on_auto_repost).pack(side="left", padx=4)
        ttk.Label(btn_row, text="使用当前选中账号执行",
                  foreground="#888").pack(side="left", padx=10)
        f9.columnconfigure(5, weight=1)

    # ============ 结果面板 ============
    def _build_results(self, parent):
        rf = ttk.Frame(parent, padding=4)
        rf.pack(fill="both", expand=True)

        # 标题行
        title_row = ttk.Frame(rf)
        title_row.pack(fill="x", pady=(0, 6))
        ttk.Label(title_row, text="🗂  结果面板",
                  font=(FONT_UI, 11, "bold")).pack(side="left")
        ttk.Label(title_row,
                  text="点 ☐ 勾选 | 列头排序 | 点 URL 复制 | 双击行打开浏览器",
                  foreground="#888", font=(FONT_UI, 9)).pack(side="left", padx=10)
        ttk.Separator(rf, orient="horizontal").pack(fill="x", pady=(0, 6))

        # 第一行：选择 + 批量操作
        bf = ttk.Frame(rf); bf.pack(fill="x", pady=(0, 3))
        self._btn(bf, "✓ 全选",   self._sel_all,   bs="secondary-outline").pack(side="left", padx=2)
        self._btn(bf, "✗ 全不选", self._sel_none,  bs="secondary-outline").pack(side="left", padx=2)
        self._btn(bf, "⇌ 反选",   self._sel_inv,   bs="secondary-outline").pack(side="left", padx=2)
        ttk.Separator(bf, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Label(bf, text="批量操作：").pack(side="left")
        self._btn(bf, "📥 采笔记+评论", self.on_bulk_detail,
                  bs="primary").pack(side="left", padx=2)
        self._btn(bf, "💗 点赞",  self.on_bulk_like,          bs="info-outline").pack(side="left", padx=2)
        self._btn(bf, "💬 评论",  self.on_bulk_comment,       bs="info-outline").pack(side="left", padx=2)
        self._btn(bf, "👤 关注",  self.on_bulk_follow_author, bs="info-outline").pack(side="left", padx=2)
        self._btn(bf, "📋 复制URL", self.on_copy_urls,        bs="secondary").pack(side="left", padx=2)

        # 第二行：导入 / 导出 / 清空
        bf2 = ttk.Frame(rf); bf2.pack(fill="x", pady=(0, 4))
        self._btn(bf2, "📂 导入URL",      self.on_import_urls,     bs="secondary").pack(side="left", padx=2)
        self._btn(bf2, "🌍 仅采IP属地",  self.on_bulk_ip_only,    bs="secondary").pack(side="left", padx=2)
        ttk.Separator(bf2, orient="vertical").pack(side="left", fill="y", padx=6)
        self._btn(bf2, "💾 导出 Excel",   self.on_export_results,  bs="success").pack(side="left", padx=2)
        self._btn(bf2, "🗑 清空结果",     self._clear_results,     bs="danger-outline").pack(side="left", padx=2)
        self.lbl_count = ttk.Label(bf2, text="共 0 条 / 已勾 0",
                                   foreground="#888", font=(FONT_UI, 10, "bold"))
        self.lbl_count.pack(side="right", padx=10)

        # ---- 过滤行（多字段筛选） ----
        ff = ttk.Frame(rf); ff.pack(fill="x", pady=(0, 4))
        ttk.Label(ff, text="🔍 关键词:").pack(side="left")
        self.e_filter = ttk.Entry(ff, width=14)
        self.e_filter.pack(side="left", padx=2)
        self.e_filter.bind("<KeyRelease>", self._apply_filter)

        ttk.Label(ff, text=" 属地:").pack(side="left", padx=(6, 0))
        self.cb_ip_filter = ttk.Combobox(ff, values=["全部"], state="readonly", width=8)
        self.cb_ip_filter.set("全部")
        self.cb_ip_filter.pack(side="left", padx=2)
        self.cb_ip_filter.bind("<<ComboboxSelected>>", self._apply_filter)

        ttk.Label(ff, text=" 类型:").pack(side="left", padx=(6, 0))
        self.cb_type_filter = ttk.Combobox(ff, values=["全部", "normal", "video"],
                                            state="readonly", width=8)
        self.cb_type_filter.set("全部")
        self.cb_type_filter.pack(side="left", padx=2)
        self.cb_type_filter.bind("<<ComboboxSelected>>", self._apply_filter)

        ttk.Label(ff, text=" 点赞≥").pack(side="left", padx=(6, 0))
        self.e_min_like = ttk.Entry(ff, width=7)
        self.e_min_like.pack(side="left", padx=2)
        self.e_min_like.bind("<KeyRelease>", self._apply_filter)

        ttk.Label(ff, text=" 评论≥").pack(side="left", padx=(6, 0))
        self.e_min_cmt = ttk.Entry(ff, width=6)
        self.e_min_cmt.pack(side="left", padx=2)
        self.e_min_cmt.bind("<KeyRelease>", self._apply_filter)

        ttk.Label(ff, text=" 收藏≥").pack(side="left", padx=(6, 0))
        self.e_min_col = ttk.Entry(ff, width=6)
        self.e_min_col.pack(side="left", padx=2)
        self.e_min_col.bind("<KeyRelease>", self._apply_filter)

        ttk.Button(ff, text="✕ 清除", command=self._clear_filter, width=7).pack(side="left", padx=(8, 2))
        ttk.Button(ff, text="✓ 勾选筛选结果", command=self._check_filtered,
                   width=14).pack(side="left", padx=2)
        ttk.Button(ff, text="📊 笔记属地分布", command=self.on_show_ip_stats).pack(side="left", padx=2)
        ttk.Button(ff, text="🗺 评论属地汇总",
                   command=self.on_show_comment_ip_stats).pack(side="left", padx=2)
        # 当前命中数显示
        self.lbl_filter_hit = ttk.Label(ff, text="", foreground="#0a7",
                                         font=(FONT_UI, 10, "bold"))
        self.lbl_filter_hit.pack(side="left", padx=8)

        # Treeview
        wrap = ttk.Frame(rf); wrap.pack(fill="both", expand=True)
        cols = ("chk", "title", "author", "ip", "likes", "comments", "collects", "type", "url")
        self.tvr = ttk.Treeview(wrap, columns=cols, show="headings", selectmode="extended")
        headers = [("chk", "☐", 40), ("title", "标题", 240), ("author", "作者", 90),
                   ("ip", "属地", 70), ("likes", "点赞", 70), ("comments", "评论", 70),
                   ("collects", "收藏", 70), ("type", "类型", 60), ("url", "URL", 320)]
        for c, label, w in headers:
            self.tvr.heading(c, text=label,
                             command=(lambda c=c: self._sort_results(c)) if c != "chk" else self._sel_invert_header)
            self.tvr.column(c, width=w, anchor="w",
                            stretch=(c in ("title", "url")), minwidth=50)
        self.tvr.tag_configure("checked", background="#1e4d2b", foreground="#fff")
        self.tvr.pack(side="left", fill="both", expand=True)
        rsb = ttk.Scrollbar(wrap, orient="vertical", command=self.tvr.yview)
        rsb.pack(side="right", fill="y")
        self.tvr.config(yscrollcommand=rsb.set)
        self.tvr.bind("<Button-1>", self._on_tvr_click)
        self.tvr.bind("<Double-1>", self._on_tvr_double)

    # ---------- 结果面板交互 ----------
    def _row_values(self, r):
        return (
            "☐",
            (r.get("title") or "")[:60],
            r.get("author", ""),
            r.get("ip_location", "") or "",
            r.get("liked_count", ""),
            r.get("comment_count", ""),
            r.get("collected_count", ""),
            r.get("type", ""),
            r.get("url", ""),
        )

    def _set_results(self, rows):
        def fill():
            self.tvr.delete(*self.tvr.get_children())
            self.results_data.clear()
            for r in rows:
                nid = r.get("note_id", "") or r.get("url", "")
                if not nid:
                    continue
                self.results_data[nid] = r
                self.tvr.insert("", "end", iid=nid, values=self._row_values(r), tags=())
            self._refresh_ip_filter_options()
            self._refresh_type_filter_options()
            self._update_count()
            self._apply_filter()
        self.root.after(0, fill)

    def _append_results(self, rows):
        def add():
            for r in rows:
                nid = r.get("note_id", "") or r.get("url", "")
                if not nid or nid in self.results_data:
                    continue
                self.results_data[nid] = r
                self.tvr.insert("", "end", iid=nid, values=self._row_values(r), tags=())
            self._update_count()
            self._apply_filter()
        self.root.after(0, add)

    def _update_row_ip(self, note_id, ip_location):
        """ 详情采集后回填 IP 属地到结果面板某行 """
        def upd():
            if note_id in self.results_data:
                self.results_data[note_id]["ip_location"] = ip_location
                if note_id in self.tvr.get_children(""):
                    self.tvr.set(note_id, "ip", ip_location or "")
                # 刷新属地下拉
                self._refresh_ip_filter_options()
        self.root.after(0, upd)

    def _on_tvr_click(self, event):
        col = self.tvr.identify_column(event.x)
        iid = self.tvr.identify_row(event.y)
        if not iid:
            return
        if col == "#1":
            self._toggle_check(iid)
        elif col == "#9":  # URL 列
            url = self.tvr.set(iid, "url")
            if url:
                self.root.clipboard_clear()
                self.root.clipboard_append(url)
                self.log("SYS", f"📋 已复制 URL: {url[:80]}")
                # 视觉反馈：闪一下选中
                self.tvr.selection_set(iid)

    def _on_tvr_double(self, event):
        iid = self.tvr.identify_row(event.y)
        if iid and iid in self.results_data:
            url = self.results_data[iid].get("url", "")
            if url:
                import webbrowser
                webbrowser.open(url)

    def _toggle_check(self, iid):
        tags = list(self.tvr.item(iid, "tags"))
        vals = list(self.tvr.item(iid, "values"))
        if "checked" in tags:
            tags.remove("checked"); vals[0] = "☐"
        else:
            tags.append("checked"); vals[0] = "☑"
        self.tvr.item(iid, values=vals, tags=tags)
        self._update_count()

    def _sel_all(self):
        for iid in self.tvr.get_children(""):
            tags = list(self.tvr.item(iid, "tags"))
            if "checked" not in tags:
                tags.append("checked")
                vals = list(self.tvr.item(iid, "values"))
                vals[0] = "☑"
                self.tvr.item(iid, values=vals, tags=tags)
        self._update_count()

    def _check_filtered(self):
        """ 一键勾选当前过滤后可见的所有行（与 _sel_all 区别：只勾筛选命中） """
        visible = set(self.tvr.get_children(""))
        n = 0
        for iid in visible:
            tags = list(self.tvr.item(iid, "tags"))
            if "checked" not in tags:
                tags.append("checked")
                vals = list(self.tvr.item(iid, "values"))
                vals[0] = "☑"
                self.tvr.item(iid, values=vals, tags=tags)
                n += 1
        self._update_count()
        if n == 0:
            self.log("SYS", "ℹ 当前可见行已全部勾选")
        else:
            self.log("SYS", f"✓ 已勾选 {n} 条筛选结果")

    def _sel_none(self):
        for iid in self.tvr.get_children(""):
            tags = list(self.tvr.item(iid, "tags"))
            if "checked" in tags:
                tags.remove("checked")
                vals = list(self.tvr.item(iid, "values"))
                vals[0] = "☐"
                self.tvr.item(iid, values=vals, tags=tags)
        self._update_count()

    def _sel_inv(self):
        for iid in self.tvr.get_children(""):
            self._toggle_check(iid)

    def _sel_invert_header(self):
        # 点击 ☐ 列头 = 反选
        self._sel_inv()

    def _clear_results(self):
        self.tvr.delete(*self.tvr.get_children())
        self.results_data.clear()
        self._refresh_ip_filter_options()
        self._refresh_type_filter_options()
        self._update_count()

    def _parse_count_filter(self, txt):
        """ 解析筛选框的数字，支持 'k' '万' 后缀 """
        s = (txt or "").strip().lower()
        if not s:
            return 0
        try:
            if s.endswith("k"):
                return int(float(s[:-1]) * 1000)
            if s.endswith("w") or s.endswith("万"):
                return int(float(s[:-1]) * 10000)
            return int(float(s))
        except Exception:
            return 0

    def _apply_filter(self, _evt=None):
        kw = self.e_filter.get().strip()
        ip_sel = self.cb_ip_filter.get() if hasattr(self, "cb_ip_filter") else "全部"
        type_sel = self.cb_type_filter.get() if hasattr(self, "cb_type_filter") else "全部"
        min_like = self._parse_count_filter(self.e_min_like.get() if hasattr(self, "e_min_like") else "")
        min_cmt = self._parse_count_filter(self.e_min_cmt.get() if hasattr(self, "e_min_cmt") else "")
        min_col = self._parse_count_filter(self.e_min_col.get() if hasattr(self, "e_min_col") else "")

        def _to_num(v):
            return parse_xhs_count(v)

        hits = 0
        for iid, r in self.results_data.items():
            match = True
            # 1) 关键词（标题/作者）
            if kw:
                blob = " ".join([str(r.get("title", "")), str(r.get("author", ""))])
                if kw not in blob:
                    match = False
            # 2) 属地下拉
            if match and ip_sel and ip_sel != "全部":
                cur = (r.get("ip_location") or "").strip()
                if ip_sel == "(未知)":
                    if cur:
                        match = False
                else:
                    if cur != ip_sel:
                        match = False
            # 3) 类型
            if match and type_sel and type_sel != "全部":
                if (r.get("type") or "") != type_sel:
                    match = False
            # 4) 数值下限
            if match and min_like > 0 and _to_num(r.get("liked_count")) < min_like:
                match = False
            if match and min_cmt > 0 and _to_num(r.get("comment_count")) < min_cmt:
                match = False
            if match and min_col > 0 and _to_num(r.get("collected_count")) < min_col:
                match = False
            try:
                if match:
                    self.tvr.reattach(iid, "", "end")
                    hits += 1
                else:
                    self.tvr.detach(iid)
            except Exception:
                pass
        # 更新命中数提示
        try:
            total = len(self.results_data)
            if hits < total:
                self.lbl_filter_hit.config(text=f"命中 {hits}/{total}")
            else:
                self.lbl_filter_hit.config(text="")
        except Exception:
            pass
        self._update_count()

    def _clear_filter(self):
        self.e_filter.delete(0, "end")
        if hasattr(self, "cb_ip_filter"):
            self.cb_ip_filter.set("全部")
        if hasattr(self, "cb_type_filter"):
            self.cb_type_filter.set("全部")
        if hasattr(self, "e_min_like"):
            self.e_min_like.delete(0, "end")
        if hasattr(self, "e_min_cmt"):
            self.e_min_cmt.delete(0, "end")
        if hasattr(self, "e_min_col"):
            self.e_min_col.delete(0, "end")
        self._apply_filter()

    def _refresh_ip_filter_options(self):
        """ 重建属地下拉列表（按出现次数排序） """
        from collections import Counter
        if not hasattr(self, "cb_ip_filter"):
            return
        cnt = Counter()
        empty = 0
        for r in self.results_data.values():
            ip = (r.get("ip_location") or "").strip()
            if ip:
                cnt[ip] += 1
            else:
                empty += 1
        opts = ["全部"]
        for ip, _ in cnt.most_common():
            opts.append(ip)
        if empty:
            opts.append("(未知)")
        cur = self.cb_ip_filter.get()
        self.cb_ip_filter["values"] = opts
        if cur not in opts:
            self.cb_ip_filter.set("全部")

    def _refresh_type_filter_options(self):
        """ 重建类型下拉列表 """
        if not hasattr(self, "cb_type_filter"):
            return
        types = set()
        for r in self.results_data.values():
            t = (r.get("type") or "").strip()
            if t:
                types.add(t)
        opts = ["全部"] + sorted(types)
        cur = self.cb_type_filter.get()
        self.cb_type_filter["values"] = opts
        if cur not in opts:
            self.cb_type_filter.set("全部")

    def on_show_ip_stats(self):
        """ 当前结果面板属地分布 """
        if not self.results_data:
            return messagebox.showinfo("提示", "结果面板为空")
        from collections import Counter
        cnt = Counter()
        for r in self.results_data.values():
            ip = (r.get("ip_location") or "").strip() or "(未知/未采详情)"
            cnt[ip] += 1
        top = cnt.most_common(30)
        dlg = tk.Toplevel(self.root)
        dlg.title("📊 属地分布")
        dlg.geometry("380x460")
        dlg.transient(self.root)
        ttk.Label(dlg, text=f"结果共 {len(self.results_data)} 条",
                  font=("", 10, "bold")).pack(pady=6)
        tv = ttk.Treeview(dlg, columns=("ip", "n"), show="headings", height=18)
        tv.heading("ip", text="属地"); tv.column("ip", width=200, anchor="w")
        tv.heading("n", text="数量"); tv.column("n", width=100, anchor="center")
        for ip, n in top:
            tv.insert("", "end", values=(ip, n))
        tv.pack(fill="both", expand=True, padx=8, pady=4)
        # 点击行 = 自动写到属地下拉
        def on_pick(_evt=None):
            sel = tv.selection()
            if sel:
                ip = tv.set(sel[0], "ip")
                if ip and not ip.startswith("("):
                    self.cb_ip_filter.set(ip)
                    self._apply_filter()
                    self.log("SYS", f"已按属地筛选: {ip}")
                    dlg.destroy()
        tv.bind("<Double-1>", on_pick)
        ttk.Label(dlg, text="双击行 = 用该属地筛选结果",
                  foreground="#888").pack(pady=4)
        ttk.Button(dlg, text="关闭", command=dlg.destroy).pack(pady=4)

    def on_show_comment_ip_stats(self):
        """ 评论作者属地汇总 - 读取 output 目录里的 note_*.xlsx，
            汇总所有评论的 ip_location 分布 """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return messagebox.showerror("缺依赖", "需要 openpyxl: pip install openpyxl")
        files = sorted(self.out_dir.glob("note_*.xlsx"))
        if not files:
            return messagebox.showinfo("提示",
                "暂无评论数据\n请先：勾选笔记 → 点 [📥 批量采笔记+评论]")
        from collections import Counter
        cnt = Counter()
        note_count = 0
        total_comments = 0
        for f in files:
            try:
                wb = load_workbook(f, read_only=True, data_only=True)
                if "评论" not in wb.sheetnames:
                    continue
                ws = wb["评论"]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = rows[0]
                try:
                    idx = headers.index("ip_location")
                except ValueError:
                    continue
                note_count += 1
                for r in rows[1:]:
                    if idx < len(r):
                        ip = (r[idx] or "")
                        if isinstance(ip, str) and ip.strip():
                            cnt[ip.strip()] += 1
                            total_comments += 1
                wb.close()
            except Exception:
                continue
        if not cnt:
            return messagebox.showinfo("提示",
                f"扫描了 {len(files)} 个文件，但没找到带属地的评论\n"
                "可能：① 评论数据较旧（重新采一次）② 评论真的没属地")
        top = cnt.most_common(50)
        dlg = tk.Toplevel(self.root)
        dlg.title("🗺 评论作者属地汇总")
        dlg.geometry("440x540")
        dlg.transient(self.root)
        ttk.Label(dlg, text=f"扫描 {len(files)} 个笔记文件 · "
                            f"含属地评论 {total_comments} 条",
                  font=(FONT_UI, 11, "bold"),
                  foreground="#F25928").pack(pady=8, padx=10, anchor="w")
        ttk.Label(dlg, text="（可帮助分析：哪些地区的人在评论你的目标笔记 → 投放/选品参考）",
                  foreground="#888").pack(pady=(0, 4), padx=10, anchor="w")
        tv = ttk.Treeview(dlg, columns=("ip", "n", "pct"),
                          show="headings", height=20)
        tv.heading("ip", text="属地"); tv.column("ip", width=120, anchor="w")
        tv.heading("n", text="评论数"); tv.column("n", width=90, anchor="center")
        tv.heading("pct", text="占比"); tv.column("pct", width=100, anchor="center")
        for ip, n in top:
            pct = f"{n * 100 / total_comments:.1f}%"
            tv.insert("", "end", values=(ip, n, pct))
        tv.pack(fill="both", expand=True, padx=10, pady=4)

        def do_export():
            from tkinter import filedialog
            p = filedialog.asksaveasfilename(
                title="导出评论属地分布",
                defaultextension=".xlsx",
                initialfile=f"评论属地汇总_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                filetypes=[("Excel", "*.xlsx")])
            if not p: return
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active
            ws.title = "评论属地分布"
            ws.append(["属地", "评论数", "占比"])
            for ip, n in cnt.most_common():
                ws.append([ip, n, f"{n*100/total_comments:.2f}%"])
            wb.save(p)
            self.log("SYS", f"评论属地汇总 → {p}")
            messagebox.showinfo("完成", f"✓ 已导出\n{p}", parent=dlg)

        bb = ttk.Frame(dlg); bb.pack(pady=8)
        ttk.Button(bb, text="💾 导出 Excel", command=do_export).pack(side="left", padx=4)
        ttk.Button(bb, text="关闭", command=dlg.destroy).pack(side="left", padx=4)

    def _update_count(self):
        total = len(self.tvr.get_children(""))
        checked = sum(1 for iid in self.tvr.get_children("")
                      if "checked" in self.tvr.item(iid, "tags"))
        self.lbl_count.config(text=f"共 {total} 条 / 已勾 {checked}")

    def _get_checked(self):
        return [self.results_data[iid] for iid in self.tvr.get_children("")
                if "checked" in self.tvr.item(iid, "tags")
                and iid in self.results_data]

    def _sort_results(self, col):
        items = [(self.tvr.set(iid, col), iid) for iid in self.tvr.get_children("")]
        asc = not self._sort_dir.get(col, False)
        self._sort_dir[col] = asc
        numeric = col in ("likes", "comments", "collects")
        if numeric:
            items.sort(key=lambda x: parse_xhs_count(x[0]), reverse=not asc)
        else:
            items.sort(key=lambda x: str(x[0]), reverse=not asc)
        for i, (_, iid) in enumerate(items):
            self.tvr.move(iid, "", i)
        labels = {"title": "标题", "author": "作者", "ip": "属地",
                  "likes": "点赞", "comments": "评论",
                  "collects": "收藏", "type": "类型", "url": "URL"}
        for c, label in labels.items():
            arrow = ("▲" if asc else "▼") if c == col else ""
            self.tvr.heading(c, text=f"{label} {arrow}")

    # ============ session 管理 ============
    def _log_cb(self, account, msg):
        self.log_q.put(f"[{time.strftime('%H:%M:%S')}][{account}] {msg}")

    def log(self, account, msg):
        self._log_cb(account, msg)

    def _get_session(self, name, create_if_missing=False):
        sess = self.sessions.get(name)
        # 清理死掉的会话
        if sess and (sess.dead or sess.start_failed):
            self.sessions.pop(name, None)
            sess = None
        if sess is None and create_if_missing:
            # 优先级：代理池绑定 > 账号独立代理
            proxy = pp.get_proxy_for_account(name) or get_proxy(name)
            sess = AccountSession(name, self._log_cb, proxy=proxy)
            sess.alert_callback = self._on_account_alert
            self.sessions[name] = sess
            if proxy:
                src = "代理池" if pp.get_proxy_for_account(name) else "独立"
                self.log(name, f"绑定代理[{src}]: {proxy[:80]}")
            else:
                self.log(name, "未配置代理（用本机 IP）")
        return sess

    def _on_account_alert(self, account, reason):
        """ 风控警报 - 弹窗 """
        def show():
            messagebox.showwarning(
                "🚨 风控警报",
                f"账号 [{account}] 触发风控保护:\n\n{reason}\n\n"
                "已自动暂停该账号所有任务。\n"
                "建议:\n"
                "  1. 检查浏览器是否出现验证码 → 手动过码\n"
                "  2. 等 30 分钟-2 小时再继续\n"
                "  3. 在账号面板右键 → 清除警报 后才能重新派任务",
            )
        self.root.after(0, show)

    def _submit_to_selected(self, fn, *args, create=True, automation_label=None):
        """ 派单到当前选中账号。
            automation_label: 不为 None 时，标记为"自动化任务"，期间禁止用户派其他任务。
        """
        name = self.acc_var.get()
        if not name:
            return messagebox.showwarning("提示", "请先选择账号")
        sess = self._get_session(name, create_if_missing=create)
        if sess is None:
            return messagebox.showwarning("提示", f"账号 [{name}] 未启动")

        # 非自动化任务派单前：如果该账号正在跑自动化，必须先暂停
        if automation_label is None and sess.automation_running:
            cur = sess.automation_label or "自动化任务"
            if not messagebox.askyesno(
                "⚠ 自动化任务运行中",
                f"账号 [{name}] 正在执行自动化任务【{cur}】\n\n"
                "现在派发新任务会与自动化抢占浏览器，可能触发风控。\n"
                "建议先暂停自动化任务再执行当前操作。\n\n"
                "→ 是: 立即停止自动化并执行当前任务\n"
                "→ 否: 取消本次操作"):
                return
            sess.stop_task()
            self.log(name, f"⏹ 已请求停止自动化 [{cur}]，等待当前步骤结束...")

        # 若浏览器还在 launch 中（首次启动），给一点时间避免任务被吃在冷启动里
        if not sess.ready_event.is_set():
            if not sess.ready_event.wait(timeout=0.5):
                self.log(name, "⏳ 浏览器仍在启动，任务已排队，启动完成后会自动执行")

        if automation_label:
            sess.submit_automation(fn, args, automation_label)
        else:
            sess.task_q.put((fn, args))

    # ============ 账号操作 ============
    def _refresh_accounts(self):
        accs = list_accounts()
        if not self.acc_var.get() or self.acc_var.get() not in accs:
            self.acc_var.set(accs[0] if accs else "")

    def on_tv_select(self, _evt=None):
        sel = self.tv.selection()
        if sel:
            self.acc_var.set(sel[0])

    def on_start_all(self):
        accs = list_accounts()
        if not accs:
            return messagebox.showinfo("提示", "没有账号")
        if not messagebox.askyesno("确认", f"启动并尝试登录全部 {len(accs)} 个账号？"):
            return
        for a in accs:
            sess = self._get_session(a, create_if_missing=True)
            sess.task_q.put((self._t_login, ()))
        self.log("SYS", f"🚀 已派发 {len(accs)} 个登录任务")

    def on_clear_alert(self):
        name = self.acc_var.get()
        sess = self.sessions.get(name)
        if not sess:
            return messagebox.showinfo("提示", f"账号 [{name}] 未启动")
        if not sess.alert:
            return messagebox.showinfo("提示", f"账号 [{name}] 未触发警报")
        if not messagebox.askyesno(
            "确认", f"清除 [{name}] 警报？\n建议先休息 30 分钟以上再继续操作"
        ):
            return
        sess.clear_alert()
        self.log("SYS", f"已清除 [{name}] 风控警报")

    def on_regen_fp(self):
        import fingerprint as _fp
        name = self.acc_var.get()
        if not name:
            return messagebox.showwarning("提示", "请先选择账号")
        if not messagebox.askyesno(
            "重置指纹",
            f"将为账号 [{name}] 重新生成浏览器指纹（UA / viewport / 硬件参数）\n"
            "✓ 适用：怀疑被关联检测、风控告警后\n"
            "⚠ 需要 [关窗口] 后重新登录才会生效"):
            return
        fp = _fp.regenerate(name)
        self.log("SYS", f"[{name}] 指纹已重置  UA={fp['user_agent'][:60]}...  "
                         f"viewport={fp['viewport']}")
        messagebox.showinfo("完成", "✓ 指纹已重置\n请关闭浏览器窗口后重新启动账号")

    def _on_save_settings_btn(self):
        self._save_settings()
        messagebox.showinfo("提示", "设置已保存")

    def on_acc_delete(self):
        name = self.acc_var.get()
        if not name:
            return messagebox.showwarning("提示", "请先选择账号")
        if not messagebox.askyesno(
            "确认删除",
            f"确定删除账号 [{name}]?\n\n"
            "将清理:\n"
            f"  - accounts/{name}.json  (cookie)\n"
            "  - _config.json 中该账号配置（代理/昵称）\n\n"
            "数据库历史和操作记录不会删除。"
        ):
            return
        # 关掉运行中的会话
        if name in self.sessions:
            try:
                self.sessions[name].shutdown()
                self.sessions.pop(name, None)
            except Exception:
                pass
        # 删 cookie 文件
        try:
            (ACCOUNTS_DIR / f"{name}.json").unlink(missing_ok=True)
        except Exception:
            pass
        # 删 _config.json 里的条目
        try:
            from scraper import load_config, save_config
            cfg = load_config()
            if name in cfg:
                cfg.pop(name, None)
                save_config(cfg)
        except Exception:
            pass
        self._refresh_accounts()
        self.log("SYS", f"已删除账号 [{name}]")

    def on_acc_new(self):
        name = simpledialog.askstring("新建账号", "账号别名（英文/数字/下划线）:", parent=self.root)
        if not name: return
        name = name.strip()
        if not name.replace("_", "").isalnum():
            return messagebox.showwarning("提示", "只能用字母数字下划线")
        p = ACCOUNTS_DIR / f"{name}.json"
        if not p.exists():
            p.write_text("{}", encoding="utf-8")
        self._refresh_accounts()
        self.acc_var.set(name)
        self.log("SYS", f"已建账号 [{name}]")

    def on_proxy_edit(self):
        name = self.acc_var.get()
        if not name:
            return messagebox.showwarning("提示", "请先选择账号")
        current = get_proxy(name)

        dlg = tk.Toplevel(self.root)
        dlg.title(f"代理配置 - {name}")
        dlg.geometry("680x360")
        dlg.transient(self.root); dlg.grab_set()

        ttk.Label(dlg, text=f"账号: {name}", font=("", 10, "bold")).pack(anchor="w", padx=12, pady=(10, 4))
        ttk.Label(dlg, text="支持两种格式（任选其一，留空=不用代理）:").pack(anchor="w", padx=12)
        ttk.Label(dlg, text="① 固定代理:  http://用户名:密码@主机:端口", foreground="#0a7").pack(anchor="w", padx=12)
        ttk.Label(dlg, text="② 代理池 API:  api:https://share.proxy.qg.net/get?key=...", foreground="#0a7").pack(anchor="w", padx=12)
        ttk.Label(dlg, text="   👆 注意前缀 api:  支持青果/芝麻等服务的提取 API", foreground="#888").pack(anchor="w", padx=12, pady=(0, 6))

        ent = ttk.Entry(dlg, width=90); ent.insert(0, current); ent.pack(padx=12, pady=4, fill="x")
        msg = tk.StringVar()
        ttk.Label(dlg, textvariable=msg, foreground="#0a7", wraplength=640, justify="left").pack(anchor="w", padx=12, pady=6)

        def do_validate():
            v = ent.get().strip()
            if not v: msg.set("（空：不使用代理）"); return True
            if v.startswith("api:"):
                msg.set("✓ 代理池模式，启动时将从 API 抽 IP"); return True
            try:
                p = parse_proxy(v)
                msg.set(f"✓ 固定代理 -> {p['server']}" + (" (含认证)" if "username" in p else ""))
                return True
            except Exception as e:
                msg.set(f"✗ {e}"); return False

        def do_test():
            v = ent.get().strip()
            if not v.startswith("api:"):
                msg.set("当前不是 api: 模式"); return
            msg.set("请求中..."); dlg.update_idletasks()
            try:
                ips = fetch_proxy_pool(v[4:].strip(), use_cache=False)
                msg.set(f"✓ 返回 {len(ips)} 个 IP，示例: {', '.join(ips[:5])}")
            except Exception as e:
                msg.set(f"✗ {e}")

        def do_save():
            if not do_validate(): return
            set_proxy(name, ent.get().strip())
            self.log("SYS", f"[{name}] 代理已保存")
            if name in self.sessions:
                messagebox.showinfo("提示", "需先 [✕ 关窗口] 再 [启动并登录] 才会应用", parent=dlg)
            dlg.destroy()

        b = ttk.Frame(dlg); b.pack(pady=10)
        ttk.Button(b, text="校验", command=do_validate).pack(side="left", padx=4)
        ttk.Button(b, text="🧪 测试代理池", command=do_test).pack(side="left", padx=4)
        ttk.Button(b, text="保存", command=do_save).pack(side="left", padx=4)
        ttk.Button(b, text="取消", command=dlg.destroy).pack(side="left", padx=4)

    def on_check_ip(self):
        self._submit_to_selected(self._t_check_ip)

    # ============ 代理 IP 池管理（#10） ============
    def on_proxy_pool(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("🌐 代理 IP 池管理")
        dlg.geometry("960x620")
        dlg.minsize(800, 500)
        dlg.transient(self.root)

        ttk.Label(dlg, text="代理 IP 池",
                  font=(FONT_UI, 13, "bold"),
                  foreground="#F25928").pack(anchor="w", padx=12, pady=(10, 2))
        ttk.Label(dlg,
                  text="每个账号绑定一个独立 IP 防关联 · 失效自动剔除 · 一键健康检测",
                  foreground="#888").pack(anchor="w", padx=12, pady=(0, 6))

        # ---- 工具栏 ----
        tb = ttk.Frame(dlg); tb.pack(fill="x", padx=12, pady=4)
        ttk.Button(tb, text="➕ 批量添加",
                   command=lambda: self._proxy_pool_add(refresh)).pack(side="left", padx=2)
        ttk.Button(tb, text="🧪 检测所有",
                   command=lambda: self._proxy_pool_check_all(refresh)).pack(side="left", padx=2)
        ttk.Button(tb, text="🗑 清理失效",
                   command=lambda: self._proxy_pool_clear_dead(refresh)).pack(side="left", padx=2)
        ttk.Separator(tb, orient="vertical").pack(side="left", fill="y", padx=8)
        ttk.Button(tb, text="🔗 自动绑定账号",
                   command=lambda: self._proxy_pool_auto_bind(refresh)).pack(side="left", padx=2)
        ttk.Button(tb, text="❌ 选中解绑",
                   command=lambda: self._proxy_pool_unbind_selected(tv, refresh)).pack(side="left", padx=2)
        ttk.Button(tb, text="🗑 删除选中",
                   command=lambda: self._proxy_pool_del_selected(tv, refresh)).pack(side="left", padx=2)

        # ---- 列表 ----
        wrap = ttk.Frame(dlg); wrap.pack(fill="both", expand=True, padx=12, pady=6)
        cols = ("url", "status", "latency", "bound", "fail", "last")
        tv = ttk.Treeview(wrap, columns=cols, show="headings", selectmode="extended")
        for c, label, w in [("url", "代理 URL", 320), ("status", "状态", 90),
                            ("latency", "延迟(ms)", 80), ("bound", "绑定账号", 100),
                            ("fail", "失败次数", 80), ("last", "上次检测", 140)]:
            tv.heading(c, text=label); tv.column(c, width=w, anchor="w")
        tv.tag_configure("ok", foreground="#0a7")
        tv.tag_configure("bad", foreground="#c33")
        tv.tag_configure("warn", foreground="#c80")
        tv.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(wrap, orient="vertical", command=tv.yview)
        sb.pack(side="right", fill="y"); tv.config(yscrollcommand=sb.set)

        # 状态栏
        status_var = tk.StringVar(value="")
        ttk.Label(dlg, textvariable=status_var,
                  foreground="#0a7").pack(anchor="w", padx=12, pady=4)

        self._proxy_pool_status = status_var

        def refresh():
            tv.delete(*tv.get_children())
            proxies = pp.list_proxies()
            ok = sum(1 for p in proxies if p["status"] == "可用")
            bound = sum(1 for p in proxies if p.get("bound_to"))
            for p in proxies:
                tag = "ok" if p["status"] == "可用" else (
                    "bad" if p["status"] in ("失效", "格式错") else "warn")
                tv.insert("", "end", iid=p["url"], tags=(tag,),
                          values=(p["url"], p["status"], p.get("latency_ms", 0),
                                  p.get("bound_to", "") or "—",
                                  p.get("fail_count", 0),
                                  p.get("last_check", "") or "—"))
            status_var.set(f"共 {len(proxies)} 条，可用 {ok}，已绑定 {bound}")
        refresh()

        ttk.Button(dlg, text="关闭", command=dlg.destroy).pack(pady=6)

    def _proxy_pool_add(self, refresh_cb):
        from tkinter import filedialog
        dlg = tk.Toplevel(self.root)
        dlg.title("批量添加代理")
        dlg.geometry("620x440")
        dlg.transient(self.root); dlg.grab_set()
        ttk.Label(dlg, text="每行一个 - 支持以下格式:",
                  font=(FONT_UI, 11, "bold")).pack(anchor="w", padx=10, pady=(8, 2))
        ttk.Label(dlg, text="✓ http://user:pwd@host:port\n"
                            "✓ http://host:port\n"
                            "✓ host:port   (自动加 http://)\n"
                            "✓ socks5://host:port",
                  foreground="#888", justify="left").pack(anchor="w", padx=10)
        txt = tk.Text(dlg, height=15, font=(FONT_MONO, 11),
                      bg="#1e2329", fg="#d4d4d4", insertbackground="#fff")
        txt.pack(fill="both", expand=True, padx=10, pady=6)
        def pick_file():
            p = filedialog.askopenfilename(filetypes=[("文本", "*.txt"), ("所有", "*.*")])
            if not p: return
            try:
                txt.insert("end", "\n" + Path(p).read_text(encoding="utf-8", errors="ignore"))
            except Exception as e:
                messagebox.showerror("错误", str(e))
        def from_pool_api():
            from tkinter import simpledialog
            api = simpledialog.askstring("代理池 API",
                                          "输入提取 API（如青果 https://share.proxy.qg.net/get?key=...）",
                                          parent=dlg)
            if not api: return
            try:
                from scraper import fetch_proxy_pool
                ips = fetch_proxy_pool(api, use_cache=False)
                txt.insert("end", "\n" + "\n".join(ips))
                messagebox.showinfo("成功", f"已拉取 {len(ips)} 个 IP，请确认后导入", parent=dlg)
            except Exception as e:
                messagebox.showerror("失败", str(e), parent=dlg)
        bb = ttk.Frame(dlg); bb.pack(pady=4)
        ttk.Button(bb, text="📁 从文件加载", command=pick_file).pack(side="left", padx=3)
        ttk.Button(bb, text="🌐 从代理 API 拉取", command=from_pool_api).pack(side="left", padx=3)
        def do_save():
            lines = [ln for ln in txt.get("1.0", "end").splitlines() if ln.strip()]
            if not lines:
                return messagebox.showwarning("提示", "请输入内容", parent=dlg)
            added, dup = pp.add_proxies(lines)
            self.log("SYS", f"代理池: 添加 {added}，重复 {dup}")
            messagebox.showinfo("完成", f"✓ 添加 {added} 条\n重复跳过 {dup} 条", parent=dlg)
            dlg.destroy()
            refresh_cb()
        bb2 = ttk.Frame(dlg); bb2.pack(pady=6)
        ttk.Button(bb2, text="✅ 保存", command=do_save).pack(side="left", padx=4)
        ttk.Button(bb2, text="取消", command=dlg.destroy).pack(side="left", padx=4)

    def _proxy_pool_check_all(self, refresh_cb):
        if not messagebox.askyesno("确认", "对所有代理进行健康检测？\n（可能需要 30~60 秒）"):
            return
        def run():
            self._proxy_pool_status.set("正在检测...")
            def prog(d, t):
                try: self._proxy_pool_status.set(f"检测中 {d}/{t}")
                except Exception: pass
            ok, total = pp.check_all(progress_cb=prog)
            self.root.after(0, refresh_cb)
            self.log("SYS", f"代理池检测完成：{ok}/{total} 可用")
        threading.Thread(target=run, daemon=True).start()

    def _proxy_pool_clear_dead(self, refresh_cb):
        if not messagebox.askyesno("确认", "删除所有失效代理（失败≥3次或不可用）？"):
            return
        n = pp.clear_dead()
        self.log("SYS", f"代理池：清理 {n} 个失效")
        refresh_cb()

    def _proxy_pool_auto_bind(self, refresh_cb):
        accs = list_accounts()
        if not accs:
            return messagebox.showwarning("提示", "没有账号")
        if not messagebox.askyesno("自动绑定",
                                    f"将给 {len(accs)} 个账号各分配一个【可用】代理\n"
                                    "（已绑定的账号会跳过）继续？"):
            return
        n = pp.auto_bind(accs)
        # 同步写入账号配置（兼容旧的 set_proxy 系统）
        for acc in accs:
            proxy = pp.get_proxy_for_account(acc)
            if proxy:
                set_proxy(acc, proxy)
        self.log("SYS", f"代理池：自动绑定 {n} 个账号")
        messagebox.showinfo("完成", f"✓ 分配 {n} 个账号")
        refresh_cb()

    def _proxy_pool_unbind_selected(self, tv, refresh_cb):
        sel = tv.selection()
        if not sel: return
        for url in sel:
            data = pp._load()
            for p in data["proxies"]:
                if p["url"] == url and p.get("bound_to"):
                    pp.unbind(p["bound_to"])
        refresh_cb()

    def _proxy_pool_del_selected(self, tv, refresh_cb):
        sel = tv.selection()
        if not sel: return
        if not messagebox.askyesno("确认", f"删除选中的 {len(sel)} 个代理？"):
            return
        for url in sel:
            pp.remove_proxy(url)
        refresh_cb()

    # ============ 图片处理工具（#9） ============
    def on_image_tool(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("🖼 图片处理工具")
        dlg.geometry("680x560")
        dlg.minsize(560, 460)
        dlg.transient(self.root)

        ttk.Label(dlg, text="图片处理 - 水印 / MD5 扰动",
                  font=(FONT_UI, 13, "bold"),
                  foreground="#F25928").pack(anchor="w", padx=12, pady=(10, 2))
        ttk.Label(dlg,
                  text="用于：发图前批量加水印、加随机噪点改变 MD5（防平台查重）",
                  foreground="#888").pack(anchor="w", padx=12, pady=(0, 6))

        # 输入文件
        in_var = tk.StringVar()
        ttk.Label(dlg, text="输入：").pack(anchor="w", padx=12, pady=(8, 2))
        rin = ttk.Frame(dlg); rin.pack(fill="x", padx=12)
        ttk.Entry(rin, textvariable=in_var).pack(side="left", fill="x", expand=True)
        def pick_in():
            from tkinter import filedialog
            paths = filedialog.askopenfilenames(
                title="选择图片（可多选）",
                filetypes=[("图片", "*.jpg *.jpeg *.png *.webp"), ("所有", "*.*")])
            if paths:
                in_var.set(";".join(paths))
        ttk.Button(rin, text="📁 选图", command=pick_in).pack(side="left", padx=4)

        # 输出目录
        out_var = tk.StringVar(value=str(self.out_dir / "processed"))
        ttk.Label(dlg, text="输出目录：").pack(anchor="w", padx=12, pady=(8, 2))
        rout = ttk.Frame(dlg); rout.pack(fill="x", padx=12)
        ttk.Entry(rout, textvariable=out_var).pack(side="left", fill="x", expand=True)
        def pick_out():
            from tkinter import filedialog
            p = filedialog.askdirectory(initialdir=out_var.get())
            if p: out_var.set(p)
        ttk.Button(rout, text="📂 选目录", command=pick_out).pack(side="left", padx=4)

        # 操作选项
        opt = ttk.LabelFrame(dlg, text="处理选项", padding=10)
        opt.pack(fill="x", padx=12, pady=10)

        v_noise = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt, text="✓ 随机噪点（改变 MD5）", variable=v_noise).grid(row=0, column=0, sticky="w", pady=2)

        v_mark = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt, text="✓ 添加水印", variable=v_mark).grid(row=1, column=0, sticky="w", pady=2)
        ttk.Label(opt, text="水印文字:").grid(row=1, column=1, sticky="e", padx=6)
        e_mark = ttk.Entry(opt, width=20); e_mark.insert(0, "@小红书博主"); e_mark.grid(row=1, column=2, sticky="w")
        ttk.Label(opt, text="位置:").grid(row=2, column=1, sticky="e", padx=6, pady=2)
        cb_pos = ttk.Combobox(opt, values=["右下", "右上", "左下", "左上", "居中"],
                              state="readonly", width=10)
        cb_pos.set("右下"); cb_pos.grid(row=2, column=2, sticky="w")
        ttk.Label(opt, text="透明度:").grid(row=3, column=1, sticky="e", padx=6)
        e_op = ttk.Entry(opt, width=10); e_op.insert(0, "180"); e_op.grid(row=3, column=2, sticky="w")
        ttk.Label(opt, text="(0~255)", foreground="#888").grid(row=3, column=3, sticky="w")

        v_crop = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt, text="✓ 微裁边（1~3 像素，改变图像哈希）",
                        variable=v_crop).grid(row=4, column=0, sticky="w", pady=2)

        v_rename = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt, text="✓ 输出随机文件名",
                        variable=v_rename).grid(row=5, column=0, sticky="w", pady=2)

        msg = tk.StringVar()
        ttk.Label(dlg, textvariable=msg, foreground="#0a7",
                  wraplength=640).pack(anchor="w", padx=12, pady=4)

        def do_run():
            paths = [p for p in in_var.get().split(";") if p.strip()]
            if not paths:
                msg.set("✗ 请先选图"); return
            out_dir = Path(out_var.get())
            try:
                out_dir.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                msg.set(f"✗ 输出目录无法创建: {e}"); return
            opts = {
                "noise": v_noise.get(),
                "watermark": v_mark.get(),
                "watermark_text": e_mark.get().strip(),
                "watermark_pos": cb_pos.get(),
                "watermark_opacity": int(e_op.get() or 180),
                "micro_crop": v_crop.get(),
                "random_name": v_rename.get(),
            }
            msg.set(f"处理中... 0/{len(paths)}"); dlg.update_idletasks()
            def run():
                ok = fail = 0
                for i, p in enumerate(paths, 1):
                    try:
                        image_proc.process(p, out_dir, opts)
                        ok += 1
                    except Exception as e:
                        fail += 1
                        self.log("IMG", f"✗ {p}: {e}")
                    self.root.after(0, lambda i=i: msg.set(
                        f"处理中... {i}/{len(paths)}"))
                self.root.after(0, lambda: msg.set(
                    f"✓ 完成 成功 {ok}  失败 {fail}  → {out_dir}"))
                self.log("IMG", f"图片处理完成 {ok}/{len(paths)} → {out_dir}")
            threading.Thread(target=run, daemon=True).start()

        bb = ttk.Frame(dlg); bb.pack(pady=10)
        ttk.Button(bb, text="🚀 开始处理", command=do_run).pack(side="left", padx=6)
        ttk.Button(bb, text="📂 打开输出目录",
                   command=lambda: os.startfile(str(Path(out_var.get())))
                   if Path(out_var.get()).exists() else None).pack(side="left", padx=6)
        ttk.Button(bb, text="关闭", command=dlg.destroy).pack(side="left", padx=6)

    def on_close_selected(self):
        name = self.acc_var.get()
        sess = self.sessions.get(name)
        if not sess:
            return messagebox.showinfo("提示", f"[{name}] 未启动")
        if not messagebox.askyesno("确认", f"关闭 [{name}] 的浏览器？"):
            return
        sess.shutdown()
        self.sessions.pop(name, None)
        self.log("SYS", f"已关闭 [{name}]")

    # ============ 日志 / 状态 ============
    def _poll_log(self):
        try:
            while True:
                m = self.log_q.get_nowait()
                self.txt.insert("end", m + "\n"); self.txt.see("end")
                if getattr(self, "log_file", None):
                    try: self.log_file.write(m + "\n")
                    except Exception: pass
        except queue.Empty: pass
        self.root.after(150, self._poll_log)

    def _poll_status(self):
        accs = list_accounts()
        existing = set(self.tv.get_children())
        for a in accs:
            meta = get_account_meta(a)
            sess = self.sessions.get(a)
            nick = (sess.nickname if sess else "") or meta.get("nickname", "") or "—"
            raw = meta.get("proxy", "")
            if not raw:
                proxy = "—"
            elif raw.startswith("api:"):
                ip = (sess.scraper.assigned_ip if sess and sess.scraper else "") or "?"
                proxy = f"🎲池:{ip}"
            else:
                proxy = "🌐固定"
            if not sess:
                login = "—"; status = "○ 未启动"
            elif sess.dead:
                login = "—"; status = "✗ 已断开"
            elif sess.start_failed:
                login = "—"; status = "✗ 启动失败"
            elif not sess.started:
                login = "—"; status = "… 启动中"
            elif sess.alert:
                login = "✓" if sess.logged_in else "✗"
                status = f"🚨 风控:{sess.alert_reason[:18]}"
            elif sess.automation_running:
                login = "✓" if sess.logged_in else "✗"
                elapsed = int(time.time() - sess.automation_started_at) if sess.automation_started_at else 0
                mm, ss = divmod(elapsed, 60)
                status = f"🤖 {sess.automation_label[:14]} {mm:02d}:{ss:02d}"
            else:
                login = "✓" if sess.logged_in else "✗"
                status = "● 执行中" if sess.busy else "○ 空闲"
            vals = (a, nick, proxy, login, status)
            if a in existing:
                self.tv.item(a, values=vals)
            else:
                self.tv.insert("", "end", iid=a, values=vals)
        for iid in existing - set(accs):
            self.tv.delete(iid)
        cur = self.acc_var.get()
        if cur and cur in accs and not self.tv.selection():
            self.tv.selection_set(cur)
        self.root.after(800, self._poll_status)

    def _ask_main(self, title, msg):
        result = []; done = threading.Event()
        def show():
            result.append(messagebox.askyesno(title, msg)); done.set()
        self.root.after(0, show); done.wait()
        return result[0]

    def _check_time_window(self):
        try:
            hs = int(self.e_hstart.get() or "10")
            he = int(self.e_hend.get() or "23")
            if not (0 <= hs <= 24 and 0 <= he <= 24 and hs < he):
                raise ValueError("invalid range")
        except ValueError:
            return self._ask_main("时间窗设置异常",
                "时间窗参数不合法，可能是误输入。是否继续？")
        h = datetime.now().hour
        if hs <= h < he:
            return True
        return self._ask_main("时间窗",
            f"当前 {h}:xx 不在 {hs}-{he} 安全窗，凌晨/深夜易触发风控。继续？")

    # ============ 任务函数（在 session 线程内执行） ============
    def _t_login(self, scraper):
        scraper.login_wait()

    def _t_check_ip(self, scraper):
        scraper.check_ip()

    def _t_search(self, scraper, kw, n):
        acc = scraper.current_account
        rows = scraper.search_notes(kw, n)
        self._set_results(rows)
        # 顺便采 IP 属地（如果用户勾了）
        if self.var_search_with_ip.get() and rows:
            self.log(acc, f"🌍 搜索完毕，开始顺采 IP 属地 共 {len(rows)} 条...")
            total = len(rows)
            ok = fail = 0
            for i, r in enumerate(rows, 1):
                self._set_progress(i - 1, total, f"🌍 [{acc}] 采IP")
                try:
                    scraper._check_stop()
                except Exception:
                    self.log(acc, "  IP 采集被中断")
                    break
                nid = r.get("note_id", "")
                url = r.get("url", "") or nid
                if r.get("ip_location"):
                    ok += 1; continue   # 已有 IP 跳过
                try:
                    ip = scraper.fetch_note_ip_only(url, timeout=8)
                    if ip:
                        ok += 1
                        r["ip_location"] = ip
                        self._update_row_ip(nid, ip)
                except Exception as e:
                    fail += 1
                    self.log(acc, f"  {i}/{total} ✗ {e}")
                scraper._sleep(random.uniform(0.8, 1.6))
            self._set_progress(0, 0)
            self.log(acc, f"🌍 IP 属地采集完成  成功 {ok}  失败 {fail}")
        # 最后导出（包含 IP 的版本）
        p = export_search(rows, self.out_dir, tag=f"search_{kw}")
        self.log(acc, f"导出 -> {p.name}")
        db.save_notes(rows)

    def _t_nurture(self, scraper, duration, like_p, collect_p):
        scraper.auto_nurture(duration, like_p, collect_p)

    def _t_publish(self, scraper, title, body, images, tags, note_type, video_path):
        scraper.publish_note(title, body, images=images, tags=tags,
                             note_type=note_type, video_path=video_path)

    # ============ 自动搬运 ============
    _RP_STYLE_BY_TYPE = {
        "风景": "治愈温柔，描述光影与氛围，第一视角分享",
        "人物": "日常生活感口吻，避免评价外貌、避免暴露隐私",
        "动物": "可爱萌宠口吻，第一人称代入，多用拟声词",
        "美食": "口水文案，描述口感与场景",
        "穿搭": "穿搭博主语气，描述材质、版型、场合",
        "其他": "自然口语，分享体验",
    }

    def on_auto_repost(self):
        kw = self.e_rp_kw.get().strip()
        if not kw:
            return messagebox.showwarning("提示", "请输入搜索关键词")
        try:
            count = int(self.e_rp_count.get() or "3")
            min_like = int(self.e_rp_min_like.get() or "1000")
            dmin = float(self.e_rp_dmin.get() or "300")
            dmax = float(self.e_rp_dmax.get() or "900")
        except ValueError:
            return messagebox.showwarning("提示", "数字参数格式错误")
        if count <= 0 or count > 20:
            return messagebox.showwarning("提示", "搬运数量建议 1-20")
        if not ai.is_enabled():
            return messagebox.showwarning(
                "提示", "未配置 DeepSeek API Key，请先到 ⚙ 设置 → AI 配置")

        rp_type = self.cb_rp_type.get()
        style = self.e_rp_style.get().strip() or self._RP_STYLE_BY_TYPE.get(rp_type, "")
        mode = self.var_rp_mode.get()

        opts = {
            "noise": self.var_rp_noise.get(),
            "micro_crop": self.var_rp_crop.get(),
            "watermark": self.var_rp_wm.get(),
            "watermark_text": self.e_rp_wm_text.get().strip(),
            "watermark_pos": "右下",
            "watermark_opacity": 180,
            "random_name": self.var_rp_rand_name.get(),
        }

        warn_msg = (f"将搜索「{kw}」抓取前 {count} 篇笔记，下载图片并用 AI 改写文案。\n\n"
                    f"模式: {'🛡 半自动（预览后人工确认发布）' if mode == 'semi' else '⚠ 全自动（自动直接发布）'}\n"
                    f"类型: {rp_type} | 文案风格: {style[:24]}\n\n"
                    "⚠ 请确认你有素材使用权或已做合规处理。继续？")
        if not messagebox.askyesno("确认搬运", warn_msg):
            return
        if mode == "auto" and not messagebox.askyesno(
            "二次确认",
            "全自动模式会跳过人工预览直接发布，风控/侵权风险显著高于半自动。\n确定吗？"):
            return

        label = f"🤖 自动搬运「{kw}」x{count}"
        self._submit_to_selected(
            self._t_auto_repost,
            kw, count, min_like, rp_type, style, opts, mode, dmin, dmax,
            automation_label=label,
        )

    def _t_auto_repost(self, scraper, kw, count, min_like, rp_type,
                        style, opts, mode, dmin, dmax):
        """ 自动搬运主流程（在 session 工作线程中跑）"""
        acc = scraper.current_account
        self.log(acc, f"🤖 开始搬运: kw={kw} count={count} type={rp_type} mode={mode}")
        # 1) 搜索（取 2~3 倍数量做候选，按点赞过滤）
        scan = max(count * 3, 20)
        try:
            rows = scraper.search_notes(kw, max_count=scan)
        except Exception as e:
            self.log(acc, f"✗ 搜索失败: {e}")
            return
        # 过滤最低点赞
        def _pl(v):
            try: return parse_xhs_count(str(v))
            except Exception:
                try: return int(v)
                except Exception: return 0
        rows = [r for r in rows if _pl(r.get("liked_count", 0)) >= min_like]
        rows = rows[:count]
        if not rows:
            self.log(acc, "✗ 没有满足条件的笔记，调低最低点赞或换关键词")
            return
        self.log(acc, f"  ✓ 命中 {len(rows)} 篇笔记，开始下载与处理")

        media_root = self.out_dir / f"repost_{acc}"
        media_root.mkdir(parents=True, exist_ok=True)
        prepared = []  # 每条: {"detail":..., "images":[路径], "ai":{title,body,tags}}
        total = len(rows)
        for i, r in enumerate(rows, 1):
            try:
                scraper._check_stop()
            except Exception:
                self.log(acc, "  ⏹ 已停止"); return
            self._set_progress(i - 1, total, f"🤖 [{acc}] 搬运")
            nid = r.get("note_id", "")
            self.log(acc, f"  [{i}/{total}] 采详情: {nid}")
            try:
                detail, _ = scraper.fetch_note_detail(r.get("url") or nid,
                                                     want_comments=False)
            except Exception as e:
                self.log(acc, f"    ✗ 详情失败: {e}")
                continue
            # 仅处理图文笔记（视频搬运风险更高，暂不支持）
            if detail.get("type") and detail["type"] != "normal":
                self.log(acc, f"    跳过非图文笔记 (type={detail.get('type')})")
                continue
            # 下载图片
            try:
                saved_folder = scraper.download_media(detail, media_root)
            except Exception as e:
                self.log(acc, f"    ✗ 下载失败: {e}")
                continue
            img_paths = [str(p) for p in saved_folder.iterdir()
                         if p.is_file() and p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")]
            if not img_paths:
                self.log(acc, "    ✗ 该笔记无可用图片，跳过")
                continue
            # 图片处理
            proc_dir = media_root / f"proc_{nid}"
            processed = []
            try:
                for p in img_paths:
                    out = image_proc.process(p, proc_dir, opts)
                    processed.append(str(out))
            except Exception as e:
                self.log(acc, f"    ✗ 图片处理失败: {e}")
                continue
            self.log(acc, f"    ✓ 图 {len(processed)} 张已处理")

            # AI 改写文案
            try:
                src_title = detail.get("title") or r.get("title") or kw
                topic = f"{kw} / 参考标题: {src_title}"
                ai_out = ai.generate_note(topic, style=style, note_type="image")
            except Exception as e:
                self.log(acc, f"    ✗ AI 文案生成失败: {e}")
                continue
            self.log(acc, f"    ✓ 文案: {ai_out.get('title','')}")

            prepared.append({
                "source_id": nid,
                "source_title": src_title,
                "images": processed,
                "ai": ai_out,
            })
            # 节流
            scraper._sleep(random.uniform(2.0, 5.0))

        self._set_progress(0, 0)
        if not prepared:
            self.log(acc, "🤖 无可发布内容")
            return
        self.log(acc, f"🤖 准备好 {len(prepared)} 条待发布")

        if mode == "auto":
            # 全自动模式：直接发布，间隔随机
            for i, item in enumerate(prepared, 1):
                try: scraper._check_stop()
                except Exception:
                    self.log(acc, "  ⏹ 已停止"); return
                self.log(acc, f"  📝 [{i}/{len(prepared)}] 自动发布: "
                              f"{item['ai'].get('title','')}")
                try:
                    scraper.publish_note(
                        item["ai"].get("title", ""),
                        item["ai"].get("body", ""),
                        images=item["images"],
                        tags=item["ai"].get("tags", []),
                        note_type="image",
                    )
                    self.log(acc, "    ✓ 已发布")
                except Exception as e:
                    self.log(acc, f"    ✗ 发布失败: {e}")
                if i < len(prepared):
                    d = random.uniform(min(dmin, dmax), max(dmin, dmax))
                    self.log(acc, f"  ⏱ 间隔 {d:.0f}s")
                    scraper._sleep(d)
            self.log(acc, "🤖 全自动搬运完成")
        else:
            # 半自动：弹预览对话框，让用户勾选并触发发布
            self.root.after(0, lambda: self._show_repost_preview(acc, prepared, dmin, dmax))

    def _show_repost_preview(self, acc, prepared, dmin, dmax):
        """ 半自动模式：展示生成结果，用户勾选要发布的条目 """
        dlg = tk.Toplevel(self.root)
        dlg.title(f"🤖 搬运预览 - {acc}  共 {len(prepared)} 条")
        dlg.geometry("760x560")
        dlg.transient(self.root)

        ttk.Label(dlg, text=f"账号 [{acc}] 待发布内容（勾选你想发的，可逐条编辑标题/正文）",
                  font=(FONT_UI, 11, "bold"),
                  foreground="#F25928").pack(anchor="w", padx=12, pady=(10, 4))

        # 滚动容器
        canvas = tk.Canvas(dlg, highlightthickness=0)
        vsb = ttk.Scrollbar(dlg, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(8, 0))
        vsb.pack(side="left", fill="y")
        inner = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        def _onconf(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _onconf)

        item_widgets = []  # [(BooleanVar, title_entry, body_text, tags_entry, item_dict)]
        for i, item in enumerate(prepared, 1):
            fr = ttk.LabelFrame(inner, text=f"#{i}  原标题: {item['source_title'][:40]}",
                                padding=6)
            fr.pack(fill="x", padx=8, pady=4)
            var = tk.BooleanVar(value=True)
            top = ttk.Frame(fr); top.pack(fill="x")
            ttk.Checkbutton(top, text="✓ 发布该条", variable=var).pack(side="left")
            ttk.Label(top, text=f"  图 {len(item['images'])} 张",
                      foreground="#0a7").pack(side="left", padx=8)

            ttk.Label(fr, text="标题:").pack(anchor="w", pady=(4, 0))
            et = ttk.Entry(fr, width=80)
            et.insert(0, item["ai"].get("title", "")); et.pack(fill="x")

            ttk.Label(fr, text="正文:").pack(anchor="w", pady=(4, 0))
            bt = tk.Text(fr, height=4, width=80, font=(FONT_UI, 10))
            bt.insert("1.0", item["ai"].get("body", "")); bt.pack(fill="x")

            ttk.Label(fr, text="标签（逗号分隔）:").pack(anchor="w", pady=(4, 0))
            tagse = ttk.Entry(fr, width=80)
            tagse.insert(0, ",".join(item["ai"].get("tags", []))); tagse.pack(fill="x")

            item_widgets.append((var, et, bt, tagse, item))

        # 底部按钮
        bb = ttk.Frame(dlg); bb.pack(side="bottom", fill="x", padx=10, pady=8)
        def do_publish():
            picked = []
            for var, et, bt, tagse, item in item_widgets:
                if not var.get(): continue
                title = et.get().strip()
                body = bt.get("1.0", "end").strip()
                tags_raw = tagse.get().strip()
                tags = [t.strip() for t in tags_raw.replace("，", ",").split(",")
                        if t.strip()]
                if not title:
                    continue
                picked.append({
                    "title": title, "body": body, "tags": tags,
                    "images": item["images"],
                })
            if not picked:
                return messagebox.showwarning("提示", "没有勾选要发布的内容", parent=dlg)
            if not messagebox.askyesno(
                "确认发布", f"将发布 {len(picked)} 条到账号 [{acc}]，"
                f"每条间隔 {int(dmin)}-{int(dmax)}s",
                parent=dlg):
                return
            dlg.destroy()
            self._submit_to_selected(
                self._t_repost_publish_batch, acc, picked, dmin, dmax,
                automation_label=f"🤖 搬运发布 x{len(picked)}",
            )
        ttk.Button(bb, text=f"📝 发布勾选项",
                   command=do_publish).pack(side="left", padx=4)
        ttk.Button(bb, text="取消", command=dlg.destroy).pack(side="right", padx=4)

    def _t_repost_publish_batch(self, scraper, acc, picked, dmin, dmax):
        """ 半自动模式批量发布（已在 session 工作线程） """
        total = len(picked)
        ok = fail = 0
        for i, item in enumerate(picked, 1):
            try: scraper._check_stop()
            except Exception:
                self.log(acc, "  ⏹ 已停止"); break
            self.log(acc, f"📝 [{i}/{total}] 发布: {item['title']}")
            try:
                scraper.publish_note(
                    item["title"], item["body"],
                    images=item["images"], tags=item["tags"],
                    note_type="image",
                )
                ok += 1
                self.log(acc, "  ✓ 已发布")
            except Exception as e:
                fail += 1
                self.log(acc, f"  ✗ 发布失败: {e}")
            if i < total:
                d = random.uniform(min(dmin, dmax), max(dmin, dmax))
                self.log(acc, f"  ⏱ 间隔 {d:.0f}s")
                scraper._sleep(d)
        self.log(acc, f"🤖 搬运发布完成  成功 {ok}  失败 {fail}")

    # ---- 笔记发布 UI ----
    def _on_pub_type_change(self):
        t = self.var_pub_type.get()
        if t == "image":
            self.lbl_pub_media.config(text="图片:")
            self.pic_frame_ref.grid()
        elif t == "video":
            self.lbl_pub_media.config(text="视频:")
            self.pic_frame_ref.grid()
        else:  # longtext
            self.lbl_pub_media.config(text="（长文无媒体）")
            # 长文也保留 listbox 但不必填

    def on_pub_ai_gen(self):
        topic = self.e_pub_topic.get().strip()
        style = self.e_pub_style.get().strip()
        if not topic:
            return messagebox.showwarning("提示", "请输入主题（如 减脂餐）")
        if not ai.is_enabled():
            return messagebox.showwarning("提示", "未配置 DeepSeek API Key，请先到 ⚙ 设置 → AI 配置")
        ntype = self.var_pub_type.get()
        # 异步调用避免卡 UI
        def work():
            try:
                self.log("SYS", f"✨ AI 生成中: {topic} ({ntype}) {style}")
                r = ai.generate_note(topic, style=style, note_type=ntype)
                def fill():
                    self.e_pub_title.delete(0, "end")
                    self.e_pub_title.insert(0, r["title"])
                    self.txt_pub_body.delete("1.0", "end")
                    self.txt_pub_body.insert("1.0", r["body"])
                    self.e_pub_tags.delete(0, "end")
                    self.e_pub_tags.insert(0, ",".join(r["tags"]))
                    self.log("SYS", f"✨ 已生成: {r['title']}")
                self.root.after(0, fill)
            except Exception as e:
                self.log("SYS", f"✗ AI 生成失败: {e}")
                self.root.after(0, lambda: messagebox.showerror("AI 生成失败", str(e)))
        threading.Thread(target=work, daemon=True).start()

    def _pub_add_imgs(self):
        from tkinter import filedialog
        t = self.var_pub_type.get()
        if t == "video":
            files = filedialog.askopenfilenames(
                filetypes=[("视频", "*.mp4 *.mov *.avi")],
            )
            # 视频只能一个
            if files:
                self.lst_pub_imgs.delete(0, "end")
                self.lst_pub_imgs.insert("end", files[0])
        else:
            files = filedialog.askopenfilenames(
                filetypes=[("图片", "*.jpg *.jpeg *.png *.webp")],
            )
            for f in files:
                self.lst_pub_imgs.insert("end", f)

    def _pub_clear_imgs(self):
        self.lst_pub_imgs.delete(0, "end")

    def _refresh_pub_accounts(self):
        for w in self._pub_acc_frame.winfo_children():
            w.destroy()
        self._pub_acc_vars.clear()
        accs = list_accounts()
        if not accs:
            ttk.Label(self._pub_acc_frame, text="(无账号)").pack(side="left")
            return
        for i, a in enumerate(accs):
            v = tk.BooleanVar(value=False)
            self._pub_acc_vars[a] = v
            cb = ttk.Checkbutton(self._pub_acc_frame, text=a, variable=v)
            cb.pack(side="left", padx=4)
            if i > 0 and i % 6 == 0:
                pass  # 简单换行不实现，多了就横滚

    def on_publish(self):
        ntype = self.var_pub_type.get()
        title = self.e_pub_title.get().strip()
        body = self.txt_pub_body.get("1.0", "end").strip()
        tags_raw = self.e_pub_tags.get().strip()
        tags = [t.strip() for t in tags_raw.replace("，", ",").split(",") if t.strip()]
        media = list(self.lst_pub_imgs.get(0, "end"))
        selected_accs = [n for n, v in self._pub_acc_vars.items() if v.get()]

        if not title:
            return messagebox.showwarning("提示", "标题必填")
        if ntype == "image" and not media:
            return messagebox.showwarning("提示", "图文模式需添加图片")
        if ntype == "video" and not media:
            return messagebox.showwarning("提示", "视频模式需添加视频文件")
        if ntype == "longtext" and len(body) < 50:
            return messagebox.showwarning("提示", "长文正文至少 50 字")
        if not selected_accs:
            return messagebox.showwarning("提示", "请勾选至少 1 个账号")

        type_name = {"image": "图文", "video": "视频", "longtext": "长文"}[ntype]
        media_desc = ""
        if ntype == "image": media_desc = f"图片: {len(media)} 张\n"
        elif ntype == "video": media_desc = f"视频: {media[0]}\n"

        if not messagebox.askyesno(
            "确认发布",
            f"类型: {type_name}\n将发布到 {len(selected_accs)} 个账号:\n  {', '.join(selected_accs)}\n\n"
            f"标题: {title}\n{media_desc}标签: {len(tags)} 个"
        ):
            return

        images = media if ntype == "image" else None
        video = media[0] if ntype == "video" and media else None

        delay_offset = 0
        for acc in selected_accs:
            sess = self._get_session(acc, create_if_missing=True)
            if sess is None: continue
            if delay_offset > 0:
                def task(scraper, t=title, b=body, im=images, tg=tags,
                         nt=ntype, vp=video, d=delay_offset):
                    scraper._sleep(d)
                    scraper.publish_note(t, b, images=im, tags=tg,
                                         note_type=nt, video_path=vp)
                sess.task_q.put((task, ()))
            else:
                sess.task_q.put((self._t_publish,
                                 (title, body, images, tags, ntype, video)))
            delay_offset += random.uniform(30, 60)
        self.log("SYS", f"📝 派发 {type_name} 笔记到 {len(selected_accs)} 个账号")

    def _t_hot(self, scraper, kw, scan, mlike, mcmt, mcol, sort_by, top):
        acc = scraper.current_account
        rows = scraper.search_hot_notes(kw, scan, mlike, mcmt, mcol, sort_by, top)
        self._set_results(rows)
        # 顺便采 IP 属地（复用搜索开关）
        if self.var_search_with_ip.get() and rows:
            self.log(acc, f"🌍 爆品采集完毕，开始顺采 IP 属地 共 {len(rows)} 条...")
            total = len(rows); ok = fail = 0
            for i, r in enumerate(rows, 1):
                self._set_progress(i - 1, total, f"🌍 [{acc}] 采IP")
                try: scraper._check_stop()
                except Exception:
                    self.log(acc, "  IP 采集被中断"); break
                nid = r.get("note_id", "")
                url = r.get("url", "") or nid
                if r.get("ip_location"):
                    ok += 1; continue
                try:
                    ip = scraper.fetch_note_ip_only(url, timeout=8)
                    if ip:
                        ok += 1
                        r["ip_location"] = ip
                        self._update_row_ip(nid, ip)
                except Exception as e:
                    fail += 1
                scraper._sleep(random.uniform(0.8, 1.6))
            self._set_progress(0, 0)
            self.log(acc, f"🌍 IP 采集完成  成功 {ok}  失败 {fail}")
        p = export_search(rows, self.out_dir, tag=f"hot_{kw}")
        self.log(acc, f"🔥 爆品 {len(rows)} 条 -> {p.name}")
        db.save_notes(rows)

    def _t_note(self, scraper, u, n, dl_media, extract_intent,
                auto_reply, reply_tpls, use_ai, dmin, dmax, dedup):
        d, c = scraper.fetch_note_detail(u, True, n)
        p = export_note(d, c, self.out_dir)
        self.log(scraper.current_account, f"导出 -> {p.name}")
        db.save_notes([d])
        # 回填属地到结果面板
        nid = d.get("note_id")
        if nid:
            self._update_row_ip(nid, d.get("ip_location", ""))
        if dl_media:
            try: scraper.download_media(d, self.out_dir)
            except Exception as e: self.log(scraper.current_account, f"媒体失败: {e}")
        intents = []
        if extract_intent and c:
            intents = extract_intent_users(c, note_meta=d)
            if intents:
                ip = export_intent_users(intents, self.out_dir, tag=d.get("note_id", ""))
                self.log(scraper.current_account, f"🎯 意向 {len(intents)} 人 -> {ip.name}")
                db.save_intent_users(intents)
        if auto_reply and intents:
            self._do_intent_replies(scraper, intents, reply_tpls, dmin, dmax, use_ai, dedup)

    def _do_intent_replies(self, scraper, intents, templates, dmin, dmax, use_ai, dedup):
        """ 给意向潜客评论自动回复（API 模式专用） """
        if not self._check_time_window():
            self.log(scraper.current_account, "时间窗外，跳过自动回复")
            return
        acc = scraper.current_account
        valid = [it for it in intents if it.get("comment_id") and it.get("note_id")]
        if not valid:
            self.log(acc, "✗ 意向用户缺少 comment_id/note_id，无法回复")
            return
        use_ai = bool(use_ai and ai.is_enabled())
        self.log(acc, f"💬 自动回复 {len(valid)} 个意向潜客 (AI={use_ai} 去重={dedup and db.is_enabled()})")
        ok = fail = skip = 0
        for i, it in enumerate(valid, 1):
            scraper._check_stop()
            cid = it["comment_id"]; nid = it["note_id"]
            if dedup and db.is_enabled() and db.has_action(acc, "reply", cid):
                skip += 1
                self.log(acc, f"  {i}/{len(valid)} 跳过（已回复过）"); continue
            template = random.choice(templates) if templates else "私我~"
            customer_comment = it.get("comment", "")
            if use_ai:
                try:
                    content = ai.reply_intent(template, customer_comment,
                                              {"title": it.get("note_title", "")})
                    self.log(acc, f"  ✨ AI 回复: {content}")
                except Exception as e:
                    content = template
                    self.log(acc, f"  AI 失败，用模板: {e}")
            else:
                content = template
            self.log(acc, f"  {i}/{len(valid)} → @{it.get('user_nickname','?')}  "
                          f"原评论:「{customer_comment[:20]}」  回复:「{content}」")
            note_url = it.get("note_url") or f"https://www.xiaohongshu.com/explore/{nid}"
            try:
                # 直接用页面模式精准回复（XHS 评论 API 需签名，未集成）
                scraper.post_reply_to_comment(note_url, cid, content)
                ok += 1
                db.record_action(acc, "reply", cid, note_url, content, "success")
            except Exception as e:
                fail += 1
                self.log(acc, f"  ✗ {e}")
                db.record_action(acc, "reply", cid, note_url, content, "failed", str(e))
            if i < len(valid):
                d = random.uniform(dmin, dmax)
                self.log(acc, f"  ⏱ {d:.0f}s")
                scraper._sleep(d)
        self.log(acc, f"自动回复完成: 成功 {ok} 失败 {fail} 跳过 {skip}")

    def _t_user(self, scraper, u, n):
        rows = scraper.fetch_user_notes(u, n)
        p = export_search(rows, self.out_dir, tag="user_notes")
        self.log(scraper.current_account, f"导出 -> {p.name}")
        self._set_results(rows)

    def _t_bulk_detail(self, scraper, urls, dl_media, extract_intent, cmt_limit,
                       auto_reply, reply_tpls, use_ai, dmin, dmax, dedup):
        acc = scraper.current_account
        self.log(acc, f"批量采笔记+评论 共 {len(urls)} 条")
        self.log(acc, f"  设置: 下载媒体={dl_media}  提取意向={extract_intent}  "
                      f"自动回复={auto_reply}  AI改写={use_ai}  去重={dedup}")
        if auto_reply and not extract_intent:
            self.log(acc, "⚠ 你勾了【自动回复】但没勾【提取意向用户】，回复不会触发！")
        if auto_reply and not reply_tpls:
            self.log(acc, "⚠ 自动回复已启用但回复模板为空")
        all_intents = []
        for i, u in enumerate(urls, 1):
            scraper._check_stop()
            self.log(acc, f"=== 详情 {i}/{len(urls)} ===")
            try:
                d, c = scraper.fetch_note_detail(u, True, cmt_limit)
                p = export_note(d, c, self.out_dir)
                self.log(acc, f"  -> {p.name}  (共 {len(c)} 条评论)")
                db.save_notes([d])
                nid_ = d.get("note_id")
                if nid_:
                    self._update_row_ip(nid_, d.get("ip_location", ""))
                if dl_media:
                    try: scraper.download_media(d, self.out_dir)
                    except Exception as e: self.log(acc, f"  媒体: {e}")
                if extract_intent and c:
                    cur_intents = extract_intent_users(c, note_meta=d)
                    all_intents.extend(cur_intents)
                    self.log(acc, f"  🎯 此笔记意向用户 {len(cur_intents)} 人")
                    if auto_reply:
                        if cur_intents:
                            self.log(acc, f"  💬 触发自动回复 {len(cur_intents)} 条 →")
                            self._do_intent_replies(scraper, cur_intents, reply_tpls,
                                                     dmin, dmax, use_ai, dedup)
                        else:
                            self.log(acc, "  💬 自动回复已启用，但此笔记无意向用户，跳过")
            except Exception as e:
                self.log(acc, f"  ✗ {e}")
            scraper._sleep(3)
        if all_intents:
            ip = export_intent_users(all_intents, self.out_dir, tag="bulk")
            self.log(acc, f"🎯 累计 {len(all_intents)} 人 -> {ip.name}")
            db.save_intent_users(all_intents)
        self.log(acc, "批量采集完成 ✓")

    def _t_bulk_like(self, scraper, rows, dmin, dmax, dedup, use_api):
        if not self._check_time_window():
            self.log(scraper.current_account, "用户取消"); return
        if not scraper.is_logged_in():
            self.log(scraper.current_account, "✗ 未登录"); return
        acc = scraper.current_account
        random.shuffle(rows)
        mode = "⚡API" if use_api else "页面"
        self.log(acc, f"💗 点赞 {len(rows)} 篇 [{mode}模式] 去重={dedup and db.is_enabled()}")
        total = len(rows)
        ok = fail = skip = 0
        for i, r in enumerate(rows, 1):
            self._set_progress(i - 1, total, f"💗 [{acc}] 点赞")
            scraper._check_stop()
            nid = r.get("note_id") or r.get("url", "")
            url = r.get("url", "")
            if dedup and db.is_enabled() and db.has_action(acc, "like", nid):
                skip += 1
                self.log(acc, f"--- {i}/{len(rows)} 跳过(已点过) ---"); continue
            self.log(acc, f"--- {i}/{len(rows)} ---")
            try:
                if use_api:
                    scraper.api_like(nid)
                else:
                    scraper.like_note(url)
                ok += 1
                self.record_action(1)  # 心跳统计
                db.record_action(acc, "like", nid, url, "", "success")
                if getattr(scraper, "session", None):
                    scraper.session.record_op("like", True)
            except Exception as e:
                fail += 1; self.log(acc, f"✗ {e}")
                db.record_action(acc, "like", nid, url, "", "failed", str(e))
                if getattr(scraper, "session", None):
                    scraper.session.record_op("like", False, str(e))
            if i < len(rows):
                # 套餐级别强制延时：用户设置不能低于服务端套餐下限
                pmin, pmax = self.plan_delay(dmin, dmax)
                d = random.uniform(pmin, pmax)
                self.log(acc, f"⏱ {d:.0f}s"); scraper._sleep(d)
        self._set_progress(0, 0)
        self.log(acc, f"完成 成功 {ok} 失败 {fail} 跳过 {skip}")

    def _t_bulk_comment(self, scraper, rows, templates, dmin, dmax,
                        need_confirm, use_ai, dedup, use_api):
        if not self._check_time_window():
            self.log(scraper.current_account, "用户取消"); return
        if not scraper.is_logged_in():
            self.log(scraper.current_account, "✗ 未登录"); return
        acc = scraper.current_account
        if use_ai and not ai.is_enabled():
            self.log(acc, "⚠ 未配置 DeepSeek，回退到模板模式")
            use_ai = False
        random.shuffle(rows)
        mode = "⚡API(edith)" if use_api else "页面"
        self.log(acc, f"💬 评论 {len(rows)} 篇 [{mode}] AI={use_ai} 去重={dedup and db.is_enabled()}")
        api_alive = use_api
        total = len(rows)
        ok = fail = skip = 0
        for i, r in enumerate(rows, 1):
            self._set_progress(i - 1, total, f"💬 [{acc}] 评论")
            scraper._check_stop()
            nid = r.get("note_id") or r.get("url", "")
            url = r.get("url", "")
            if dedup and db.is_enabled() and db.has_action(acc, "comment", nid):
                skip += 1
                self.log(acc, f"--- {i}/{len(rows)} 跳过(已评过) ---"); continue
            template = random.choice(templates)
            if use_ai:
                try:
                    content = ai.rewrite(template, r)
                    self.log(acc, f"✨ AI: {content}")
                except Exception as e:
                    content = template
                    self.log(acc, f"AI 失败，用模板: {e}")
            else:
                content = template
            self.log(acc, f"--- {i}/{len(rows)} ---")
            self.log(acc, f"目标: {url[:80]}")
            self.log(acc, f"评论: {content}")
            if need_confirm:
                if not self._ask_main(f"[{acc}] 确认",
                                       f"目标:\n{url}\n\n评论:\n{content}\n\n发送？"):
                    self.log(acc, "用户跳过"); continue
            try:
                if api_alive:
                    try:
                        scraper.api_comment(nid, content, note_url=url)
                    except Exception as e_api:
                        self.log(acc, f"⚠ API 失败（406=签名缺失），本批后续走快速页面模式: {e_api}")
                        api_alive = False
                        scraper.post_comment_fast(url, content)
                else:
                    scraper.post_comment_fast(url, content)
                ok += 1
                self.record_action(1)
                db.record_action(acc, "comment", nid, url, content, "success")
                if getattr(scraper, "session", None):
                    scraper.session.record_op("comment", True)
            except Exception as e:
                fail += 1; self.log(acc, f"✗ {e}")
                db.record_action(acc, "comment", nid, url, content, "failed", str(e))
                if getattr(scraper, "session", None):
                    scraper.session.record_op("comment", False, str(e))
            if i < len(rows):
                pmin, pmax = self.plan_delay(dmin, dmax)
                d = random.uniform(pmin, pmax)
                self.log(acc, f"⏱ {d:.0f}s"); scraper._sleep(d)
        self._set_progress(0, 0)
        self.log(acc, f"完成 成功 {ok} 失败 {fail} 跳过 {skip}")

    def _t_bulk_follow(self, scraper, items, dmin, dmax, dedup, use_api):
        """ items: [(user_id, user_url), ...] """
        if not self._check_time_window():
            self.log(scraper.current_account, "用户取消"); return
        if not scraper.is_logged_in():
            self.log(scraper.current_account, "✗ 未登录"); return
        acc = scraper.current_account
        random.shuffle(items)
        mode = "⚡API" if use_api else "页面"
        self.log(acc, f"👤 关注 {len(items)} 人 [{mode}] 去重={dedup and db.is_enabled()}")
        total = len(items)
        ok = fail = skip = 0
        for i, (uid, url) in enumerate(items, 1):
            self._set_progress(i - 1, total, f"👤 [{acc}] 关注")
            scraper._check_stop()
            if dedup and db.is_enabled() and uid and db.has_action(acc, "follow", uid):
                skip += 1
                self.log(acc, f"--- {i}/{len(items)} 跳过(已关注过) ---"); continue
            self.log(acc, f"--- {i}/{len(items)} ---")
            try:
                if use_api:
                    if not uid:
                        raise RuntimeError("API 模式需要 user_id")
                    scraper.api_follow(uid)
                else:
                    scraper.follow_user(url)
                ok += 1
                self.record_action(1)
                db.record_action(acc, "follow", uid or url, url, "", "success")
                if getattr(scraper, "session", None):
                    scraper.session.record_op("follow", True)
            except Exception as e:
                fail += 1; self.log(acc, f"✗ {e}")
                db.record_action(acc, "follow", uid or url, url, "", "failed", str(e))
                if getattr(scraper, "session", None):
                    scraper.session.record_op("follow", False, str(e))
            if i < len(items):
                pmin, pmax = self.plan_delay(dmin, dmax)
                d = random.uniform(pmin, pmax)
                self.log(acc, f"⏱ {d:.0f}s"); scraper._sleep(d)
        self._set_progress(0, 0)
        self.log(acc, f"完成 成功 {ok} 失败 {fail} 跳过 {skip}")

    # ============ 顶部按钮回调 ============
    def on_login(self): self._submit_to_selected(self._t_login)

    def on_stop_selected(self):
        sess = self.sessions.get(self.acc_var.get())
        if sess: sess.stop_task(); self.log(sess.account, "请求停止...")

    def on_stop_all(self):
        for s in self.sessions.values(): s.stop_task()
        self.log("SYS", "已请求停止全部")

    def on_search(self):
        kw = self.e_kw.get().strip()
        if not kw: return messagebox.showwarning("提示", "请输入关键词")
        self._submit_to_selected(self._t_search, kw, int(self.e_count.get() or "20"))

    def on_note(self):
        u = self.e_note.get().strip()
        if not u: return messagebox.showwarning("提示", "请输入 URL/ID")
        try:
            dmin = float(self.e_dmin.get() or "60")
            dmax = float(self.e_dmax.get() or "150")
        except ValueError:
            dmin, dmax = 60, 150
        reply_tpls = [ln.strip() for ln in self.txt_reply.get("1.0", "end").splitlines() if ln.strip()]
        self._submit_to_selected(
            self._t_note, u, int(self.e_cmt.get() or "100"),
            self.var_dl_media.get(), self.var_extract_intent.get(),
            self.var_auto_reply.get(), reply_tpls, self.var_use_ai.get(),
            dmin, dmax, self.var_dedup.get())

    def on_user(self):
        u = self.e_user.get().strip()
        if not u: return messagebox.showwarning("提示", "请输入 URL/ID")
        self._submit_to_selected(self._t_user, u, int(self.e_ucount.get() or "30"))

    def on_nurture(self):
        try:
            d = int(self.e_nur_dur.get() or "30")
            lp = float(self.e_nur_like.get() or "15") / 100
            cp = float(self.e_nur_col.get() or "5") / 100
        except ValueError:
            return messagebox.showwarning("提示", "参数须为数字")
        self._submit_to_selected(self._t_nurture, d, lp, cp,
                                 automation_label=f"🌱 自动养号 {d}min")

    def on_hot(self):
        kw = self.e_hkw.get().strip()
        if not kw: return messagebox.showwarning("提示", "请输入关键词")
        try:
            scan = int(self.e_hscan.get() or "100")
            mlike = int(self.e_hmin_like.get() or "0")
            mcmt = int(self.e_hmin_cmt.get() or "0")
            mcol = int(self.e_hmin_col.get() or "0")
            top = int(self.e_htop.get() or "30")
        except ValueError:
            return messagebox.showwarning("提示", "数字格式错误")
        sort_map = {"按点赞": "liked_count", "按评论": "comment_count", "按收藏": "collected_count"}
        self._submit_to_selected(self._t_hot, kw, scan, mlike, mcmt, mcol,
                                 sort_map.get(self.cb_hsort.get(), "liked_count"), top)

    # ============ 批量操作（基于结果面板勾选） ============
    def _checked_urls(self):
        rows = self._get_checked()
        return [r.get("url", "") for r in rows if r.get("url")]

    def _checked_author_urls(self):
        rows = self._get_checked()
        urls = []
        for r in rows:
            aid = r.get("author_id", "")
            if aid:
                urls.append(f"https://www.xiaohongshu.com/user/profile/{aid}")
        return urls

    def on_bulk_detail(self):
        urls = self._checked_urls()
        if not urls:
            return messagebox.showwarning("提示", "请先勾选目标")
        if not messagebox.askyesno("确认", f"批量采详情 {len(urls)} 条？\n（若已勾 ✨ 自动回复，会边采边回复潜客）"):
            return
        try: cn = int(self.e_cmt.get() or "100")
        except ValueError: cn = 100
        try:
            dmin = float(self.e_dmin.get() or "60")
            dmax = float(self.e_dmax.get() or "150")
        except ValueError:
            dmin, dmax = 60, 150
        reply_tpls = [ln.strip() for ln in self.txt_reply.get("1.0", "end").splitlines() if ln.strip()]
        self._submit_to_selected(
            self._t_bulk_detail, urls,
            self.var_dl_media.get(), self.var_extract_intent.get(), cn,
            self.var_auto_reply.get(), reply_tpls, self.var_use_ai.get(),
            dmin, dmax, self.var_dedup.get())

    def on_bulk_like(self):
        rows = self._get_checked()
        if not rows:
            return messagebox.showwarning("提示", "请先勾选目标")
        if not messagebox.askyesno("确认", f"批量点赞 {len(rows)} 篇？"):
            return
        try:
            dmin = float(self.e_fmin.get() or "30")
            dmax = float(self.e_fmax.get() or "90")
        except ValueError:
            return messagebox.showwarning("提示", "互动间隔参数错误（⑤ 互动参数）")
        self._submit_to_selected(self._t_bulk_like, rows, dmin, dmax,
                                 self.var_dedup.get(), self.var_api_mode.get())

    def on_bulk_comment(self):
        rows = self._get_checked()
        if not rows:
            return messagebox.showwarning("提示", "请先勾选目标")
        tpls = [ln.strip() for ln in self.txt_templates.get("1.0", "end").splitlines() if ln.strip()]
        if not tpls:
            return messagebox.showwarning("提示", "请到 ④ 评论模板 填模板")
        try:
            dmin = float(self.e_dmin.get() or "60")
            dmax = float(self.e_dmax.get() or "150")
        except ValueError:
            return messagebox.showwarning("提示", "评论间隔错误（④）")
        if dmin < 30 and not messagebox.askyesno("风险", "间隔<30s，风控高，继续？"):
            return
        if not messagebox.askyesno("确认", f"批量评论 {len(rows)} 篇？"):
            return
        self._submit_to_selected(self._t_bulk_comment, rows, tpls, dmin, dmax,
                                 self.var_confirm.get(), self.var_use_ai.get(),
                                 self.var_dedup.get(), self.var_api_mode.get())

    def on_bulk_follow_author(self):
        rows = self._get_checked()
        items = []
        for r in rows:
            aid = r.get("author_id", "")
            if aid:
                items.append((aid, f"https://www.xiaohongshu.com/user/profile/{aid}"))
        if not items:
            return messagebox.showwarning("提示", "勾选的笔记里没有作者ID")
        if not messagebox.askyesno("确认", f"批量关注 {len(items)} 位作者？"):
            return
        try:
            dmin = float(self.e_fmin.get() or "30")
            dmax = float(self.e_fmax.get() or "90")
        except ValueError:
            return messagebox.showwarning("提示", "互动参数错误（⑤）")
        self._submit_to_selected(self._t_bulk_follow, items, dmin, dmax,
                                 self.var_dedup.get(), self.var_api_mode.get())

    def on_copy_urls(self):
        urls = self._checked_urls()
        if not urls:
            return messagebox.showwarning("提示", "请先勾选")
        text = "\n".join(urls)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.log("SYS", f"已复制 {len(urls)} 个 URL 到剪贴板")

    # ============ 导入 URL 列表（#7） ============
    def on_import_urls(self):
        from tkinter import filedialog
        dlg = tk.Toplevel(self.root)
        dlg.title("📂 导入笔记 URL 列表")
        dlg.geometry("680x520")
        dlg.transient(self.root); dlg.grab_set()
        ttk.Label(dlg, text="支持粘贴或导入 TXT/CSV — 每行一个 URL 或笔记ID",
                  font=(FONT_UI, 11, "bold"),
                  foreground="#F25928").pack(anchor="w", padx=12, pady=(10, 4))
        ttk.Label(dlg,
                  text="✓ 支持完整 URL：https://www.xiaohongshu.com/explore/68xxxxxx?xsec_token=...\n"
                       "✓ 支持纯笔记 ID：68aff92d000000001c010eeb\n"
                       "✓ 自动跳过空行、注释 (#开头)、重复项\n"
                       "✓ 导入后会出现在结果面板，可勾选执行批量操作",
                  foreground="#888", justify="left").pack(anchor="w", padx=12)

        tools = ttk.Frame(dlg); tools.pack(fill="x", padx=12, pady=8)
        def pick_file():
            p = filedialog.askopenfilename(
                title="选择 TXT/CSV 文件",
                filetypes=[("文本文件", "*.txt"), ("CSV", "*.csv"), ("所有文件", "*.*")])
            if not p: return
            try:
                content = Path(p).read_text(encoding="utf-8", errors="ignore")
                txt.delete("1.0", "end"); txt.insert("1.0", content)
                lbl.config(text=f"已加载 {p}", foreground="#0a7")
            except Exception as e:
                lbl.config(text=f"读取失败: {e}", foreground="#c33")
        ttk.Button(tools, text="📁 选择文件", command=pick_file).pack(side="left", padx=4)
        ttk.Button(tools, text="📋 从剪贴板粘贴",
                   command=lambda: (txt.delete("1.0", "end"),
                                    txt.insert("1.0", dlg.clipboard_get()))
                   ).pack(side="left", padx=4)
        ttk.Button(tools, text="🗑 清空",
                   command=lambda: txt.delete("1.0", "end")).pack(side="left", padx=4)

        txt = tk.Text(dlg, height=15, font=(FONT_MONO, 11),
                      bg="#1e2329", fg="#d4d4d4", insertbackground="#fff")
        txt.pack(fill="both", expand=True, padx=12, pady=(0, 6))

        lbl = ttk.Label(dlg, text="", foreground="#888")
        lbl.pack(anchor="w", padx=12)

        def do_import():
            raw = txt.get("1.0", "end").strip()
            if not raw:
                lbl.config(text="✗ 内容为空", foreground="#c33"); return
            import re as _re
            existing = set(self.results_data.keys())
            added = dup = bad = 0
            rows = []
            for line in raw.splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                # 提取 note_id + xsec_token
                m = _re.search(r"/(?:explore|discovery/item)/([0-9a-zA-Z]+)", line)
                if m:
                    nid = m.group(1)
                    url = line if line.startswith("http") else f"https://www.xiaohongshu.com/explore/{nid}"
                elif _re.fullmatch(r"[0-9a-zA-Z]{16,32}", line):
                    nid = line
                    url = f"https://www.xiaohongshu.com/explore/{nid}"
                else:
                    bad += 1; continue
                if nid in existing or nid in {r.get("note_id") for r in rows}:
                    dup += 1; continue
                rows.append({
                    "note_id": nid,
                    "title": "(待采集)", "type": "", "author": "",
                    "author_id": "", "liked_count": "", "collected_count": "",
                    "comment_count": "", "ip_location": "",
                    "url": url,
                })
                added += 1
            if not rows:
                lbl.config(text=f"✗ 没有新URL（重复 {dup}，无效 {bad}）",
                           foreground="#c33"); return
            # 追加到结果面板
            for r in rows:
                nid = r["note_id"]
                self.results_data[nid] = r
                self.tvr.insert("", "end", iid=nid, values=self._row_values(r))
            self._refresh_ip_filter_options()
            self._refresh_type_filter_options()
            self._update_count()
            self.log("SYS", f"📂 导入 {added} 条URL（重复 {dup}，无效 {bad}）")
            messagebox.showinfo("导入完成",
                                f"✓ 成功导入 {added} 条\n重复跳过 {dup} 条\n无效行 {bad} 条")
            dlg.destroy()

        bb = ttk.Frame(dlg); bb.pack(pady=8)
        ttk.Button(bb, text="✅ 导入到结果面板", command=do_import).pack(side="left", padx=6)
        ttk.Button(bb, text="取消", command=dlg.destroy).pack(side="left", padx=6)

    # ============ 仅采 IP 属地（快速） ============
    def on_bulk_ip_only(self):
        rows = self._get_checked()
        if not rows:
            return messagebox.showwarning("提示", "请先勾选目标")
        if not messagebox.askyesno("确认", f"仅采 IP 属地 {len(rows)} 条？\n"
                                          "（不抓评论/媒体，比 [批量采笔记+评论] 快 5-10×）"):
            return
        self._submit_to_selected(self._t_bulk_ip_only, rows)

    def _t_bulk_ip_only(self, scraper, rows):
        acc = scraper.current_account
        self.log(acc, f"🌍 仅采 IP 属地 共 {len(rows)} 条")
        total = len(rows)
        ok = fail = 0
        for i, r in enumerate(rows, 1):
            self._set_progress(i - 1, total, f"🌍 [{acc}] 仅采IP")
            scraper._check_stop()
            url = r.get("url") or r.get("note_id", "")
            nid = r.get("note_id") or url
            try:
                ip = scraper.fetch_note_ip_only(url)
                if ip:
                    ok += 1
                    self._update_row_ip(nid, ip)
                    self.log(acc, f"  {i}/{total} ✓ {ip}")
                else:
                    fail += 1
                    self.log(acc, f"  {i}/{total} ✗ 未取到 IP")
            except Exception as e:
                fail += 1
                self.log(acc, f"  {i}/{total} ✗ {e}")
            scraper._sleep(random.uniform(1.5, 3))
        self._set_progress(0, 0)
        self.log(acc, f"完成 成功 {ok} 失败 {fail}")

    def on_export_results(self):
        from tkinter import filedialog
        import shutil
        rows = list(self.results_data.values())
        if not rows:
            return messagebox.showwarning("提示", "结果为空")
        # 让用户选择保存路径
        default_name = f"小红书结果_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="导出 Excel 到...",
            defaultextension=".xlsx",
            initialfile=default_name,
            initialdir=str(self.out_dir),
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not save_path:
            return  # 用户取消
        # 先生成到 out_dir，再移动到用户指定位置（保留原 exporter 接口）
        try:
            tmp = export_search(rows, self.out_dir, tag="manual")
            target = Path(save_path)
            if tmp.resolve() != target.resolve():
                shutil.move(str(tmp), str(target))
            self.log("SYS", f"✓ 已导出 -> {target}")
            if messagebox.askyesno("导出成功", f"已保存到:\n{target}\n\n是否打开所在文件夹?"):
                try:
                    os.startfile(str(target.parent))
                except Exception:
                    pass
        except Exception as e:
            messagebox.showerror("导出失败", str(e))
            self.log("SYS", f"✗ 导出失败: {e}")

    def open_out(self):
        os.startfile(self.out_dir)

    def on_close(self):
        # 任务运行中警告
        busy_sess = [n for n, s in self.sessions.items() if s.busy]
        if busy_sess:
            if not messagebox.askyesno(
                "确认退出",
                f"以下账号还在执行任务：{', '.join(busy_sess)}\n"
                "强制关闭可能导致评论/点赞中断。仍然退出？"
            ):
                return
        # 保存设置
        try:
            self._save_settings()
        except Exception:
            pass
        # 关日志文件
        try:
            if getattr(self, "log_file", None):
                self.log_file.write(f"\n=== {time.strftime('%H:%M:%S')} 退出 ===\n")
                self.log_file.close()
        except Exception:
            pass
        # 关闭 sessions — 等每个 session 把 storage_state 写盘后再退出，
        # 否则 daemon 线程会被 root.destroy() 一并杀掉，下次启动小红书就要重新登录
        sess_list = list(self.sessions.values())
        for s in sess_list:
            try: s.shutdown()  # 先全部发停止信号（并行收尾，节省时间）
            except Exception: pass
        for s in sess_list:
            try: s.shutdown(join_timeout=6.0)  # 再逐个 join 等其退出
            except Exception: pass
        self.root.destroy()


def _set_app_user_model_id():
    """ Windows: 让任务栏图标用 Tk 自定义图标，而不是 python.exe 默认图标 """
    if os.name != "nt":
        return
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            "qiqi.collector.v2"
        )
    except Exception:
        pass


# ttkbootstrap 可选主题列表（按 light/dark 分组）
THEMES_LIGHT = ["cosmo", "flatly", "litera", "minty", "lumen", "sandstone",
                "yeti", "pulse", "united", "morph", "journal", "simplex", "cerculean"]
THEMES_DARK  = ["darkly", "superhero", "solar", "cyborg", "vapor"]
DEFAULT_THEME = "cosmo"   # 默认 Bootstrap 浅色风格


def _apply_fonts(root):
    """ 统一设置全局字体（与主题无关）"""
    try:
        from tkinter import font as _fnt
        for name in ("TkDefaultFont", "TkTextFont", "TkMenuFont",
                     "TkHeadingFont", "TkCaptionFont", "TkSmallCaptionFont",
                     "TkIconFont", "TkTooltipFont"):
            try: _fnt.nametofont(name).configure(family=FONT_UI, size=11)
            except Exception: pass
        try: _fnt.nametofont("TkFixedFont").configure(family=FONT_MONO, size=11)
        except Exception: pass
    except Exception: pass


def _apply_ttkb_styles(style):
    """ 在 ttkbootstrap Style 基础上叠加业务定制 """
    try:
        style.configure(".", font=(FONT_UI, 11))
        style.configure("TButton",      padding=(10, 6))
        style.configure("TEntry",       padding=5)
        style.configure("TCombobox",    padding=4)
        style.configure("TNotebook.Tab", padding=(16, 7),
                        font=(FONT_UI, 11, "bold"))
        style.configure("TLabelframe.Label", font=(FONT_UI, 11, "bold"))
        style.configure("Treeview",     rowheight=30, font=(FONT_UI, 10))
        style.configure("Treeview.Heading", font=(FONT_UI, 10, "bold"),
                        padding=(6, 7))
    except Exception: pass


def _apply_ui_theme(root, theme_name=None):
    """ 仅在使用旧版 tk.Tk() 根窗口时调用（fallback）"""
    _apply_fonts(root)
    try:
        import ttkbootstrap as ttkb
        if not theme_name:
            try:
                import settings_mgr
                theme_name = settings_mgr.get("ui_theme", DEFAULT_THEME)
            except Exception:
                theme_name = DEFAULT_THEME
        if theme_name not in THEMES_LIGHT and theme_name not in THEMES_DARK:
            theme_name = DEFAULT_THEME
        style = ttkb.Style(theme=theme_name)
        _apply_ttkb_styles(style)
        try:
            root.configure(bg=style.colors.bg)
        except Exception: pass
        root._current_theme = theme_name
        return style
    except ImportError:
        style = ttk.Style()
        try: style.theme_use("vista")
        except Exception:
            try: style.theme_use("clam")
            except Exception: pass
        style.configure(".", font=(FONT_UI, 11))
        style.configure("TButton", padding=(10, 6))
        style.configure("TLabelframe.Label",
                        font=(FONT_UI, 11, "bold"), foreground="#F25928")
        style.configure("Treeview", rowheight=30)
        style.map("Treeview",
                  background=[("selected", "#F25928")],
                  foreground=[("selected", "white")])
        root.configure(bg="#f5f7fa")
        return style
    except Exception: pass


def switch_theme(root, theme_name):
    """ 运行时切换主题 + 保存设置 """
    try:
        import ttkbootstrap as ttkb
        style = ttkb.Style(theme=theme_name)
        _apply_ttkb_styles(style)
        try: root.configure(bg=style.colors.bg)
        except Exception: pass
        root._current_theme = theme_name
        import settings_mgr
        settings_mgr.set_value("ui_theme", theme_name)
    except Exception as e:
        print(f"切换主题失败: {e}")


def main():
    _set_app_user_model_id()
    # 1) 先授权（独立窗口，不创建主 Tk）
    ok, data, err = lm.check_status()
    if not ok:
        activated = activation_ui.show_activation(parent=None, err_msg=err)
        if not activated:
            return
        ok, data, err = lm.check_status()
        if not ok:
            r = tk.Tk(); r.withdraw()
            messagebox.showerror("授权失败", err or "无法验证授权")
            r.destroy()
            return

    # 2) 授权通过，再启主窗（优先用 ttkbootstrap.Window 获得完整主题效果）
    try:
        import ttkbootstrap as ttkb
        import settings_mgr
        theme_name = settings_mgr.get("ui_theme", DEFAULT_THEME)
        if theme_name not in THEMES_LIGHT and theme_name not in THEMES_DARK:
            theme_name = DEFAULT_THEME
        root = ttkb.Window(themename=theme_name, scaling=1.0)
        root._current_theme = theme_name
        _apply_fonts(root)
        _apply_ttkb_styles(ttkb.Style())
    except Exception:
        root = tk.Tk()
        _apply_ui_theme(root)
    App(root, license_data=data)
    root.mainloop()


if __name__ == "__main__":
    main()
