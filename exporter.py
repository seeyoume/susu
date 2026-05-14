"""Excel 导出"""
from datetime import datetime
from openpyxl import Workbook


def _ts():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def export_search(rows, out_dir, tag="search"):
    wb = Workbook()
    ws = wb.active
    ws.title = "结果"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([str(r.get(h, "")) for h in headers])
    else:
        ws.append(["（无数据）"])
    path = out_dir / f"{tag}_{_ts()}.xlsx"
    wb.save(path)
    return path


def export_intent_users(rows, out_dir, tag=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "高意向用户"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([str(r.get(h, "")) for h in headers])
    else:
        ws.append(["（无匹配）"])
    safe_tag = (tag or "intent").replace("/", "_")[:30]
    path = out_dir / f"intent_{safe_tag}_{_ts()}.xlsx"
    wb.save(path)
    return path


def export_note(detail, comments, out_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "笔记详情"
    ws.append(["字段", "内容"])
    for k, v in detail.items():
        ws.append([k, str(v)])
    ws2 = wb.create_sheet("评论")
    if comments:
        headers = list(comments[0].keys())
        ws2.append(headers)
        for c in comments:
            ws2.append([str(c.get(h, "")) for h in headers])
    else:
        ws2.append(["（无评论）"])
    nid = detail.get("note_id", "x")
    path = out_dir / f"note_{nid}_{_ts()}.xlsx"
    wb.save(path)
    return path
